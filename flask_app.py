from flask import Flask, render_template, jsonify, request, session, redirect, url_for, flash
from flask_cors import CORS
from functools import wraps
from datetime import datetime, timezone, timedelta
from werkzeug.utils import secure_filename

# IST Timezone (UTC+5:30)
IST = timezone(timedelta(hours=5, minutes=30))

def get_ist_now():
    """Get current datetime in IST"""
    return datetime.now(IST)
import os
import json
import time
import re
import requests
import base64
import uuid
import qrcode
import threading
import tempfile
import shutil
import hmac
import hashlib
import logging
import io
from html import escape as html_escape
from io import BytesIO
from flask_mail import Mail, Message
from config import ALLOWED_EMAIL_DOMAINS

# API keys are now loaded from club_info.json (editable in admin panel)
# These are helper functions to get current API config
def get_api_config():
    """Get API configuration from club_info.json"""
    club_info, _, _, _ = load_data()
    return club_info.get('api_config', {})

def get_groq_api_key():
    return get_api_config().get('GROQ_API_KEY', '')

def get_groq_model():
    return get_api_config().get('GROQ_MODEL', 'llama-3.1-8b-instant')

def get_razorpay_keys():
    config = get_api_config()
    return config.get('RAZORPAY_KEY_ID', ''), config.get('RAZORPAY_KEY_SECRET', '')

# Cached events context for chatbot (updated when events change)
_events_context_cache = None

def update_events_context_cache(events_list=None):
    """Update the cached events context string for chatbot"""
    global _events_context_cache
    if events_list is None:
        events_list, _ = load_events_file()
    _events_context_cache = "\n".join([
        f"- {e.get('name')}: {e.get('description', 'No description')} | Date: {e.get('date')} | Status: {e.get('status')} | Location: {e.get('location')}"
        for e in events_list
    ])
    return _events_context_cache

def get_events_context():
    """Get cached events context, building it if needed"""
    global _events_context_cache
    if _events_context_cache is None:
        update_events_context_cache()
    return _events_context_cache

# Configure logging instead of print statements
logging.basicConfig(level=logging.INFO)
logger = logging.getLogger(__name__)

# Thread lock for file operations to prevent race conditions
_file_locks = {}
_file_locks_lock = threading.Lock()
_MAX_FILE_LOCKS = 100  # Prevent unbounded memory growth

def get_file_lock(filepath):
    """Get or create a lock for a specific file"""
    with _file_locks_lock:
        if filepath not in _file_locks:
            # Clean up old locks if we have too many (simple LRU-style cleanup)
            if len(_file_locks) >= _MAX_FILE_LOCKS:
                # Remove locks that aren't currently held
                to_remove = []
                for path, lock in list(_file_locks.items()):
                    if not lock.locked():
                        to_remove.append(path)
                        if len(_file_locks) - len(to_remove) < _MAX_FILE_LOCKS // 2:
                            break
                for path in to_remove:
                    del _file_locks[path]
            _file_locks[filepath] = threading.Lock()
        return _file_locks[filepath]

def safe_json_read(filepath):
    """Safely read JSON file with locking"""
    lock = get_file_lock(filepath)
    with lock:
        if not os.path.exists(filepath):
            return []
        try:
            with open(filepath, 'r', encoding='utf-8') as f:
                return json.load(f)
        except (json.JSONDecodeError, IOError) as e:
            logger.error(f"Failed to read JSON from {filepath}: {e}")
            # Try to recover from backup if exists
            backup_path = filepath + '.backup'
            if os.path.exists(backup_path):
                logger.info(f"Attempting to recover from backup: {backup_path}")
                try:
                    with open(backup_path, 'r', encoding='utf-8') as f:
                        return json.load(f)
                except:
                    pass
            return []

def _write_json_no_lock(filepath, data):
    """Internal: Write JSON without acquiring lock (caller must hold lock)"""
    os.makedirs(os.path.dirname(filepath), exist_ok=True)
    
    # Create backup of existing file
    if os.path.exists(filepath):
        backup_path = filepath + '.backup'
        try:
            shutil.copy2(filepath, backup_path)
        except Exception as e:
            logger.warning(f"Could not create backup: {e}")
    
    # Write to temp file first, then atomic rename
    dir_name = os.path.dirname(filepath)
    fd, temp_path = tempfile.mkstemp(suffix='.json', dir=dir_name)
    try:
        with os.fdopen(fd, 'w', encoding='utf-8') as f:
            json.dump(data, f, indent=4, ensure_ascii=False)
        
        # Atomic rename/replace
        if os.name == 'nt':
            if os.path.exists(filepath):
                os.replace(temp_path, filepath)
            else:
                os.rename(temp_path, filepath)
        else:
            os.rename(temp_path, filepath)
        
        logger.debug(f"Successfully wrote JSON to {filepath}")
        return True
    except Exception as e:
        if os.path.exists(temp_path):
            os.remove(temp_path)
        raise e

def safe_json_write(filepath, data):
    """
    Safely write JSON file with atomic write and backup.
    Uses a temp file + rename approach to prevent corruption.
    """
    lock = get_file_lock(filepath)
    with lock:
        return _write_json_no_lock(filepath, data)

def atomic_add_registration(filepath, new_registration, unique_check_fn=None):
    """
    Atomically add a registration to a JSON file.
    This ensures read-check-write happens in a single lock to prevent race conditions.
    
    Args:
        filepath: Path to the registrations JSON file
        new_registration: The registration dict to add
        unique_check_fn: Optional function(registrations, new_reg) -> error_msg or None
                        Returns error message if duplicate found, None if OK
    
    Returns:
        (success: bool, error_msg: str or None, registrations: list)
    """
    lock = get_file_lock(filepath)
    with lock:
        # Read existing registrations inside the lock
        registrations = []
        if os.path.exists(filepath):
            try:
                with open(filepath, 'r', encoding='utf-8') as f:
                    registrations = json.load(f)
            except (json.JSONDecodeError, IOError) as e:
                logger.error(f"Failed to read JSON from {filepath}: {e}")
                registrations = []
        
        # Check for duplicates if check function provided
        if unique_check_fn:
            error_msg = unique_check_fn(registrations, new_registration)
            if error_msg:
                return (False, error_msg, registrations)
        
        # Assign sequential ID
        new_registration['id'] = len(registrations) + 1
        
        # Append and write - all within the same lock
        registrations.append(new_registration)
        
        try:
            _write_json_no_lock(filepath, registrations)
            return (True, None, registrations)
        except Exception as e:
            logger.error(f"Failed to save registration: {e}")
            return (False, f"Failed to save registration: {str(e)}", registrations)

# Get the absolute path of the directory containing this file (AICC/)
BASE_DIR = os.path.dirname(os.path.abspath(__file__))
# All data, templates, and static folders are in the same AICC directory
PROJECT_ROOT = BASE_DIR

# Function to ensure all required folders and files exist
def initialize_app_structure():
    """Create all necessary folders and files if they don't exist"""
    
    # Create necessary directories in PROJECT_ROOT
    directories = [
        os.path.join(PROJECT_ROOT, 'data'),
        os.path.join(PROJECT_ROOT, 'static'),
        os.path.join(PROJECT_ROOT, 'static/uploads'),
        os.path.join(PROJECT_ROOT, 'static/css'),
        os.path.join(PROJECT_ROOT, 'static/js'),
        os.path.join(PROJECT_ROOT, 'static/img'),
        os.path.join(PROJECT_ROOT, 'static/img/members'),
        os.path.join(PROJECT_ROOT, 'static/img/life'),
        os.path.join(PROJECT_ROOT, 'static/img/poster'),
        os.path.join(PROJECT_ROOT, 'templates'),
        os.path.join(PROJECT_ROOT, 'templates/admin'),
    ]
    
    for directory in directories:
        os.makedirs(directory, exist_ok=True)
    
    # Create default JSON files if they don't exist
    data_files = {
        'data/club_info.json': {
            "name": "AI and Cybersecurity Club",
            "short_name": "AICC",
            "tagline": "Innovate. Secure. Excel.",
            "description": "The AI and Cybersecurity Club is a student-driven community dedicated to exploring cutting-edge technologies in artificial intelligence and cybersecurity.",
            "college": "Thiagarajar College of Engineering",
            "department": "Department of Computer Science and Engineering",
            "address": "Madurai, Tamil Nadu, India",
            "logo": "/static/img/aicc-logo.webp"
        },
        'data/events.json': {"next_id": 1, "events": []},
        'data/members.json': [],
        'data/gallery.json': []
    }
    
    for file_path, default_content in data_files.items():
        full_path = os.path.join(PROJECT_ROOT, file_path)
        if not os.path.exists(full_path):
            with open(full_path, 'w') as f:
                json.dump(default_content, f, indent=4)

# Initialize app structure on startup
initialize_app_structure()

app = Flask(__name__, 
            template_folder=os.path.join(PROJECT_ROOT, 'templates'),
            static_folder=os.path.join(PROJECT_ROOT, 'static'))
# SECURITY: Use environment variable for secret key in production
app.secret_key = os.environ.get('FLASK_SECRET_KEY', 'your-secret-key-change-this-in-production')

# CORS: Allow all origins on all endpoints
CORS(app)
app.config['UPLOAD_FOLDER'] = os.path.join(PROJECT_ROOT, 'static/uploads')
app.config['MAX_CONTENT_LENGTH'] = 16 * 1024 * 1024  # 16MB max file size
app.config['SEND_FILE_MAX_AGE_DEFAULT'] = 0  # Disable caching in development
ALLOWED_EXTENSIONS = {'png', 'jpg', 'jpeg', 'gif', 'webp'}

# Initialize Flask-Mail (will be configured from club_info.json)
mail = Mail(app)

def configure_mail():
    """Configure Flask-Mail from club_info.json"""
    email_config = CLUB_INFO.get('email_config', {})
    if email_config:
        app.config['MAIL_SERVER'] = email_config.get('MAIL_SERVER', 'smtp.gmail.com')
        app.config['MAIL_PORT'] = email_config.get('MAIL_PORT', 587)
        app.config['MAIL_USE_TLS'] = email_config.get('MAIL_USE_TLS', True)
        app.config['MAIL_USERNAME'] = email_config.get('MAIL_USERNAME', '')
        app.config['MAIL_PASSWORD'] = email_config.get('MAIL_PASSWORD', '')
        app.config['MAIL_DEFAULT_SENDER'] = email_config.get('MAIL_DEFAULT_SENDER', '')
        mail.init_app(app)

# Function to load data from JSON files
def load_data():
    """Reload all data from JSON files"""
    data_dir = os.path.join(PROJECT_ROOT, 'data')
    with open(os.path.join(data_dir, 'club_info.json'), 'r') as f:
        club_info = json.load(f)
    with open(os.path.join(data_dir, 'events.json'), 'r') as f:
        events_data = json.load(f)
        # Handle both old array format and new object format
        if isinstance(events_data, list):
            # Migrate old format: convert array to object with next_id
            max_id = max([e.get('id', 0) for e in events_data], default=0)
            events = events_data
            # Save migrated format
            with open(os.path.join(data_dir, 'events.json'), 'w') as fw:
                json.dump({"next_id": max_id + 1, "events": events}, fw, indent=4)
        else:
            events = events_data.get('events', [])
    with open(os.path.join(data_dir, 'members.json'), 'r') as f:
        members = json.load(f)
    with open(os.path.join(data_dir, 'gallery.json'), 'r') as f:
        gallery = json.load(f)
    return club_info, events, members, gallery

# Load initial data
CLUB_INFO, EVENTS, MEMBERS, GALLERY = load_data()

# Configure mail with loaded data
configure_mail()

def load_events_file():
    """Load events.json and return (events_list, next_id)"""
    events_file = os.path.join(PROJECT_ROOT, 'data/events.json')
    with open(events_file, 'r') as f:
        events_data = json.load(f)
    
    if isinstance(events_data, list):
        # Old format - migrate
        events = events_data
        next_id = max([e.get('id', 0) for e in events], default=0) + 1
    else:
        events = events_data.get('events', [])
        next_id = events_data.get('next_id', 1)
    
    return events, next_id

def save_events_file(events, next_id):
    """Save events list with next_id to events.json"""
    events_file = os.path.join(PROJECT_ROOT, 'data/events.json')
    with open(events_file, 'w') as f:
        json.dump({"next_id": next_id, "events": events}, f, indent=4)
    # Update chatbot context cache
    update_events_context_cache(events)

# Add cache-busting filter
@app.template_filter('cache_bust')
def cache_bust_filter(url):
    """Add timestamp to URL for cache busting"""
    if url and '?' not in url:
        return f"{url}?v={int(time.time())}"
    return url

# Make cache_bust available in all templates
app.jinja_env.filters['cache_bust'] = cache_bust_filter

# Make datetime.now available in templates
app.jinja_env.globals['now'] = datetime.now

# Admin credentials - loaded from environment variables with insecure defaults
# SECURITY: Set ADMIN_USERNAME and ADMIN_PASSWORD environment variables in production
ADMIN_USERNAME = os.environ.get('ADMIN_USERNAME', 'admin')
ADMIN_PASSWORD = os.environ.get('ADMIN_PASSWORD', 'password')

# Helper function for file uploads
def allowed_file(filename):
    return '.' in filename and filename.rsplit('.', 1)[1].lower() in ALLOWED_EXTENSIONS

def sort_members_by_role(members, role_hierarchy, year_hierarchy):
    """Sort members by predefined role hierarchy and year (descending)"""
    def get_sort_key(member):
        role = member.get('role', '')
        year = member.get('year', '')
        
        # Get role index
        try:
            role_index = role_hierarchy.index(role)
        except ValueError:
            # If role not in hierarchy, put at the end
            role_index = len(role_hierarchy)
        
        # Get year index (negative for descending order)
        try:
            year_index = -year_hierarchy.index(year)  # Negative for descending
        except ValueError:
            # If year not in hierarchy, put at the end
            year_index = 0
        
        return (role_index, year_index)
    
    return sorted(members, key=get_sort_key)

def slugify(value):
    """Create a URL-safe slug from text"""
    value = (value or '').strip().lower()
    value = re.sub(r'[^a-z0-9]+', '-', value)
    return value.strip('-') or 'event'

def delete_old_image(image_path):
    """Delete old image file if it exists in uploads folder"""
    if image_path and '/static/uploads/' in image_path:
        # Extract filename from path
        filename = image_path.split('/static/uploads/')[-1]
        filepath = os.path.join(app.config['UPLOAD_FOLDER'], filename)
        try:
            if os.path.exists(filepath):
                os.remove(filepath)
        except Exception as e:
            pass

def generate_qr_code(data_string):
    """Generate QR code and return as base64 encoded string"""
    try:
        qr = qrcode.QRCode(
            version=1,
            error_correction=qrcode.constants.ERROR_CORRECT_L,
            box_size=10,
            border=4,
        )
        qr.add_data(data_string)
        qr.make(fit=True)
        
        img = qr.make_image(fill_color="black", back_color="white")
        
        # Convert to base64
        buffer = BytesIO()
        img.save(buffer, format='PNG')
        buffer.seek(0)
        img_base64 = base64.b64encode(buffer.getvalue()).decode()
        
        return img_base64
    except Exception as e:
        logger.error(f"QR code generation error: {e}")
        return None

def send_registration_email(email, registration_id, qr_code_base64, event_name, registration_data):
    """Send registration confirmation email with QR code"""
    try:
        configure_mail()  # Reconfigure mail in case settings changed
        
        msg = Message(
            subject=f'Registration Confirmation - {event_name}',
            recipients=[email]
        )
        
        # Create HTML email body with CID reference for QR code
        # Escape user-provided data to prevent XSS
        safe_name = html_escape(registration_data.get('name', 'Participant'))
        safe_event_name = html_escape(event_name)
        safe_registration_id = html_escape(str(registration_id))
        safe_club_name = html_escape(CLUB_INFO.get('name', 'AI Coding Club'))
        safe_college = html_escape(CLUB_INFO.get('college', ''))
        
        html_body = f"""
        <html>
            <body style="font-family: Arial, sans-serif; max-width: 600px; margin: 0 auto; padding: 20px;">
                <div style="background: linear-gradient(135deg, #667eea 0%, #764ba2 100%); padding: 30px; text-align: center; border-radius: 10px 10px 0 0;">
                    <h1 style="color: white; margin: 0;">Registration Successful!</h1>
                </div>
                
                <div style="background: #f8f9fa; padding: 30px; border-radius: 0 0 10px 10px;">
                    <p style="font-size: 16px; color: #333;">Dear {safe_name},</p>
                    
                    <p style="font-size: 14px; color: #555;">
                        Thank you for registering for <strong>{safe_event_name}</strong>!
                    </p>
                    
                    <div style="background: white; padding: 20px; border-radius: 8px; margin: 20px 0; border-left: 4px solid #667eea;">
                        <h3 style="color: #667eea; margin-top: 0;">Registration ID:</h3>
                        <p style="font-size: 20px; font-weight: bold; color: #333; margin: 10px 0;">{safe_registration_id}</p>
                    </div>
                    
                    <p style="font-size: 14px; color: #555;">
                        Please save this QR code. You may need to present it at the event:
                    </p>
                    
                    <div style="text-align: center; margin: 20px 0;">
                        <img src="cid:qrcode" alt="QR Code" style="max-width: 250px; border: 2px solid #ddd; padding: 10px; background: white; border-radius: 8px;"/>
                    </div>
                    
                    <p style="font-size: 14px; color: #555;">
                        We look forward to seeing you at the event!
                    </p>
                    
                    <hr style="border: none; border-top: 1px solid #ddd; margin: 30px 0;">
                    
                    <p style="font-size: 12px; color: #999; text-align: center;">
                        This is an automated email. Please do not reply.<br>
                        {safe_club_name} | {safe_college}
                    </p>
                </div>
            </body>
        </html>
        """
        
        msg.html = html_body
        
        # Attach QR code as inline image with Content-ID
        if qr_code_base64:
            qr_image_data = base64.b64decode(qr_code_base64)
            msg.attach(
                filename='qrcode.png',
                content_type='image/png',
                data=qr_image_data,
                disposition='inline',
                headers={'Content-ID': '<qrcode>'}
            )
        
        mail.send(msg)
        return True
    except Exception as e:
        logger.error(f"Email sending error: {e}")
        return False

def create_razorpay_order(order_id, amount, customer_name, customer_email, customer_phone, return_url):
    """Create a Razorpay payment order"""
    try:
        # Razorpay API endpoint
        url = "https://api.razorpay.com/v1/orders"
        
        # Basic Auth using key_id and key_secret from config
        key_id, key_secret = get_razorpay_keys()
        auth = (key_id, key_secret)
        
        headers = {
            "Content-Type": "application/json"
        }
        
        # Convert amount to paise (Razorpay uses smallest currency unit)
        amount_in_paise = int(float(amount) * 100)
        
        # Razorpay order payload
        payload = {
            "amount": amount_in_paise,
            "currency": "INR",
            "receipt": str(order_id),
            "notes": {
                "customer_name": customer_name or "Guest",
                "customer_email": customer_email or "",
                "customer_phone": customer_phone or ""
            }
        }
        
        response = requests.post(url, json=payload, headers=headers, auth=auth)
        
        if response.status_code == 200:
            razorpay_response = response.json()
            # Return in format expected by our code
            return {
                "order_id": razorpay_response.get("id"),
                "amount": amount,
                "currency": razorpay_response.get("currency"),
                "receipt": razorpay_response.get("receipt")
            }
        else:
            # Return error details
            error_data = {'error': 'Razorpay API error', 'status_code': response.status_code}
            try:
                error_data['details'] = response.json()
            except:
                error_data['details'] = response.text
            return error_data
            
    except requests.exceptions.RequestException as e:
        return {'error': 'Network error', 'details': str(e)}
    except Exception as e:
        return {'error': 'Unexpected error', 'details': str(e)}

@app.route('/')
def home():
    """Home page with hero section and registration deadline"""
    # Reload data to get latest changes
    global CLUB_INFO, EVENTS, MEMBERS, GALLERY
    CLUB_INFO, EVENTS, MEMBERS, GALLERY = load_data()
    
    # Filter out hidden events
    visible_events = [e for e in EVENTS if e.get('show_in_events', True)]
    
    # Sort events: with register_link first, then by status (upcoming first)
    sorted_events = sorted(visible_events, key=lambda x: (
        not bool(x.get('register_link')),  # Events with register_link first
        x.get('status') != 'upcoming',     # Then upcoming events
        x.get('status') == 'completed'     # Then completed
    ))
    
    # Find the next event with an active registration deadline (sorted by earliest deadline)
    next_deadline_event = None
    valid_deadline_events = []
    
    for event in visible_events:
        # Only consider upcoming events with registration deadlines
        if event.get('status') == 'upcoming':
            if event.get('registration_deadline') and event.get('register_link'):
                deadline_date = event['registration_deadline']['date']
                try:
                    # Try multiple date formats
                    deadline = None
                    for date_format in ['%Y-%m-%d', '%B %d, %Y']:
                        try:
                            deadline = datetime.strptime(deadline_date, date_format)
                            break
                        except ValueError:
                            continue
                    
                    # Using IST for comparison
                    if deadline and deadline.date() >= get_ist_now().date():
                        valid_deadline_events.append((deadline, event))
                except Exception as e:
                    pass
    
    # Sort by earliest deadline and pick the first one
    if valid_deadline_events:
        valid_deadline_events.sort(key=lambda x: x[0])
        next_deadline_event = valid_deadline_events[0][1]
    
    return render_template('index.html', 
                         club_info=CLUB_INFO, 
                         events=sorted_events[:3],  # Show only top 3 events on home
                         contact=CLUB_INFO,
                         next_deadline_event=next_deadline_event)

@app.route('/about')
def about():
    """About page"""
    global CLUB_INFO, EVENTS, MEMBERS, GALLERY
    CLUB_INFO, EVENTS, MEMBERS, GALLERY = load_data()
    return render_template('about.html', 
                         club_info=CLUB_INFO,
                         contact=CLUB_INFO)

@app.route('/events')
def events():
    """Events page showing all events"""
    global CLUB_INFO, EVENTS, MEMBERS, GALLERY
    CLUB_INFO, EVENTS, MEMBERS, GALLERY = load_data()
    # Filter out hidden events, then sort
    visible_events = [e for e in EVENTS if e.get('show_in_events', True)]
    sorted_events = sorted(visible_events, key=lambda x: (
        not bool(x.get('register_link')),  # Events with register_link first
        x.get('status') != 'upcoming',     # Then upcoming events
        x.get('status') == 'completed'     # Then completed
    ))
    return render_template('events.html', 
                         events=sorted_events,
                         club_info=CLUB_INFO,
                         contact=CLUB_INFO)

@app.route('/api/chatbot', methods=['POST'])
def chatbot_api():
    """Chatbot API endpoint using Groq"""
    try:
        data = request.get_json()
        user_message = data.get('message', '').strip()
        
        if not user_message:
            return jsonify({'error': 'Message is required'}), 400
        
        # Get cached events context
        events_context = get_events_context()
        
        # Get contact details from club_info
        club_info, _, _, _ = load_data()
        
        # Build contact context
        faculty_contacts = "\n".join([
            f"  - {f.get('name')}: {f.get('phone')}" 
            for f in club_info.get('faculty_coordinators', [])
        ])
        secretary_contacts = "\n".join([
            f"  - {s.get('name')}: {s.get('phone')}" 
            for s in club_info.get('secretaries', [])
        ])
        
        system_prompt = f"""You are a helpful assistant for AI Coding Club (AICC) at Kongu Engineering College. You help users with information about club events, registrations, and general queries.

About AICC:
- AI Coding Club empowers students to learn, build, and innovate in AI and software development
- We conduct workshops, hackathons, and collaborative projects
- Open to all students interested in AI, coding, and technology
- Department: {club_info.get('department', 'Department of AI')}
- College: {club_info.get('college', 'Kongu Engineering College')}

Contact Information:
- Email: {club_info.get('email', 'kecaicodingclub@gmail.com')}
- Instagram: {club_info.get('instagram', '')}
- LinkedIn: {club_info.get('linkedin', '')}
- Address: {club_info.get('address', '')}

Faculty Coordinators:
{faculty_contacts}

Student Secretaries:
{secretary_contacts}

Current Events:
{events_context}

Website Features for Users:
- Home Page: Overview of the club, recent updates, and quick links
- Events Page: Browse all upcoming and past events, filter by status
- Event Details: Click any event to see full details, schedule, and requirements
- Registration: Click "Register Now" on upcoming events to sign up (some events may require payment via Razorpay)
- Attendance Check: After attending an event, check your attendance status at /attendance-check using your email or registration ID
- Gallery: View photos from past events and activities
- Members: See our team - coordinators, leads, and members
- About: Learn more about AICC's mission and activities

How to Register for Events:
1. Go to Events page and find the event
2. Click "View Details" then "Register Now"
3. Fill in your details (name, email, roll number, etc.)
4. If payment is required, complete payment via Razorpay
5. You'll receive a confirmation email with QR code

Guidelines:
- Be friendly and concise
- If asked about events, provide relevant details from the events list
- For contact queries, provide the relevant contact details
- For registration queries, guide users step by step
- If you don't know something specific, say so politely
- Keep responses brief (2-3 sentences max unless more detail is needed)
- IMPORTANT: Never use markdown formatting. No tables, no bold (**), no italics (*), no headers (#), no code blocks. Use plain text only with simple line breaks."""
        
        # Call Groq API
        response = requests.post(
            'https://api.groq.com/openai/v1/chat/completions',
            headers={
                'Authorization': f'Bearer {get_groq_api_key()}',
                'Content-Type': 'application/json'
            },
            json={
                'model': get_groq_model(),
                'messages': [
                    {'role': 'system', 'content': system_prompt},
                    {'role': 'user', 'content': user_message}
                ],
                'max_tokens': 300,
                'temperature': 0.7
            },
            timeout=30
        )
        
        if response.status_code == 200:
            result = response.json()
            bot_reply = result['choices'][0]['message']['content']
            return jsonify({'reply': bot_reply})
        else:
            logger.error(f"Groq API error: {response.status_code} - {response.text}")
            return jsonify({'error': 'Failed to get response from AI'}), 500
            
    except requests.exceptions.Timeout:
        return jsonify({'error': 'Request timed out. Please try again.'}), 504
    except Exception as e:
        logger.error(f"Chatbot error: {str(e)}")
        return jsonify({'error': 'An error occurred'}), 500

@app.route('/events/<int:event_id>')
def event_detail(event_id):
    """Individual event detail page"""
    global CLUB_INFO, EVENTS, MEMBERS, GALLERY
    CLUB_INFO, EVENTS, MEMBERS, GALLERY = load_data()
    event = next((e for e in EVENTS if e.get('id') == event_id), None)
    if not event:
        return render_template('404.html'), 404
    return render_template('event_detail.html',
                         event=event,
                         club_info=CLUB_INFO,
                         contact=CLUB_INFO)

@app.route('/events/<int:event_id>/register')
def event_register(event_id):
    """Event registration form page"""
    global CLUB_INFO, EVENTS, MEMBERS, GALLERY
    CLUB_INFO, EVENTS, MEMBERS, GALLERY = load_data()
    
    event = next((e for e in EVENTS if e.get('id') == event_id), None)
    if not event:
        return render_template('404.html'), 404
    
    # Check if event has internal registration
    if event.get('registration_type') != 'internal':
        flash('This event uses external registration.', 'info')
        return redirect(url_for('event_detail', event_id=event_id))
    
    # Check if registration is allowed (admin toggle)
    registration_closed = event.get('allow_registration') == False
    
    # Check if registration deadline has passed (using IST)
    deadline_passed = False
    deadline_info = event.get('registration_deadline')
    if deadline_info and deadline_info.get('date'):
        try:
            deadline_date = deadline_info['date']
            if deadline_date.upper() != 'TBA':
                deadline = datetime.strptime(deadline_date, '%Y-%m-%d')
                # Registration closes after the deadline day (deadline day is last day to register)
                # Using IST for comparison
                if deadline.date() < get_ist_now().date():
                    deadline_passed = True
        except ValueError:
            pass
    
    # Load form templates
    templates_file = os.path.join(PROJECT_ROOT, 'data', 'form_templates.json')
    try:
        with open(templates_file, 'r') as f:
            templates = json.load(f)
        
        # Find the template for this event
        template = next((t for t in templates if t.get('id') == event.get('template_id')), None)
        if template and not template.get('active'):
            template = None
        
        # Generate submit endpoint based on event slug
        if template and not registration_closed and not deadline_passed:
            event_slug = slugify(event.get('name'))
            submit_endpoint = f"/api/register/{event_slug}"
        else:
            submit_endpoint = None
        
        return render_template('register_form.html',
                             event=event,
                             form=template,
                             submit_endpoint=submit_endpoint,
                             registration_closed=registration_closed,
                             deadline_passed=deadline_passed,
                             club_info=CLUB_INFO,
                             contact=CLUB_INFO)
    except Exception as e:
        flash('Registration form not available at this time.', 'error')
        return redirect(url_for('event_detail', event_id=event_id))

@app.route('/api/register/<event_slug>', methods=['POST'])
def api_register_event(event_slug):
    """API endpoint to handle event registration submissions"""
    try:
        data = request.get_json()
        
        # Validate required fields
        if not data:
            return jsonify({'error': 'No data provided'}), 400
        
        # ENSURE SUBMITTER EMAIL IS ALWAYS REQUIRED
        submitter_email = data.get('submitter_email', '').strip()
        if not submitter_email:
            return jsonify({
                'error': 'Submitter email is required',
                'details': 'Please provide your email address'
            }), 400
        
        # Validate submitter email format
        email_pattern = r'^[a-zA-Z0-9._%+-]+@[a-zA-Z0-9.-]+\.[a-zA-Z]{2,}$'
        if not re.match(email_pattern, submitter_email):
            return jsonify({
                'error': 'Invalid email format',
                'details': 'Please provide a valid email address'
            }), 400
        
        # Validate email domain
        email_domain = submitter_email.split('@')[1].lower()
        if email_domain not in ALLOWED_EMAIL_DOMAINS:
            allowed_domains_str = ', '.join(ALLOWED_EMAIL_DOMAINS)
            return jsonify({
                'error': f'Email domain not allowed. Please use one of: {allowed_domains_str}'
            }), 400
        
        # Validate form if template_id provided
        template_id = data.get('template_id')
        template_definition = None
        if template_id is not None:
            try:
                template_id_int = int(template_id)
            except (TypeError, ValueError):
                return jsonify({'error': 'Invalid template id'}), 400

            templates_file = os.path.join(PROJECT_ROOT, 'data', 'form_templates.json')
            if os.path.exists(templates_file):
                with open(templates_file, 'r') as f:
                    templates = json.load(f)
                template_definition = next((t for t in templates if t.get('id') == template_id_int), None)

        if template_definition:
            if not template_definition.get('active', False):
                return jsonify({'error': 'Registration form is inactive'}), 400

            # NEW: Validate participants (if participant-based template)
            min_participants = template_definition.get('min_participants', 1)
            max_participants = template_definition.get('max_participants', 1)
            
            if min_participants and max_participants:
                num_participants = int(data.get('num_participants', min_participants))
                
                # Validate participant count
                if num_participants < min_participants or num_participants > max_participants:
                    return jsonify({
                        'error': f'Number of participants must be between {min_participants} and {max_participants}'
                    }), 400
                
                # Collect and validate participant data
                participants = []
                for i in range(1, num_participants + 1):
                    participant_name = data.get(f'participant_{i}_name', '').strip()
                    participant_roll = data.get(f'participant_{i}_roll', '').strip()
                    participant_email = data.get(f'participant_{i}_email', '').strip()
                    
                    if not participant_name:
                        return jsonify({
                            'error': f'Participant {i} name is required'
                        }), 400
                    
                    if not participant_roll:
                        return jsonify({
                            'error': f'Participant {i} roll number is required'
                        }), 400
                    
                    if not participant_email:
                        return jsonify({
                            'error': f'Participant {i} email is required'
                        }), 400
                    
                    # Validate email format
                    if not re.match(email_pattern, participant_email):
                        return jsonify({
                            'error': f'Participant {i} has invalid email format'
                        }), 400
                    
                    # Validate email domain
                    p_email_domain = participant_email.split('@')[1].lower()
                    if p_email_domain not in ALLOWED_EMAIL_DOMAINS:
                        return jsonify({
                            'error': f'Participant {i} email domain not allowed. Please use one of: {", ".join(ALLOWED_EMAIL_DOMAINS)}'
                        }), 400
                    
                    participants.append({
                        'name': participant_name,
                        'roll_no': participant_roll,
                        'email': participant_email
                    })
                
                # Store participants array in registration data
                data['participants'] = participants
                data['num_participants'] = num_participants
            
            # Validate custom fields (new structure)
            custom_fields = template_definition.get('custom_fields', [])
            missing_fields = []
            for field in custom_fields:
                if field.get('required'):
                    field_name = field.get('name')
                    if not field_name or not str(data.get(field_name, '')).strip():
                        missing_fields.append(field.get('label') or field_name)
            
            # Also check legacy 'fields' structure for backward compatibility
            legacy_fields = template_definition.get('fields', [])
            for field in legacy_fields:
                if field.get('required'):
                    field_name = field.get('name')
                    if not field_name or not str(data.get(field_name, '')).strip():
                        missing_fields.append(field.get('label') or field_name)

            if missing_fields:
                return jsonify({
                    'error': 'Missing required fields',
                    'missing': missing_fields
                }), 400
            
            # Validate email fields
            for field in template_definition.get('fields', []):
                if field.get('type') == 'email':
                    field_name = field.get('name')
                    email_value = data.get(field_name, '').strip()
                    
                    if email_value:
                        # Basic email format validation
                        email_pattern = r'^[a-zA-Z0-9._%+-]+@[a-zA-Z0-9.-]+\.[a-zA-Z]{2,}$'
                        if not re.match(email_pattern, email_value):
                            return jsonify({
                                'error': f'Invalid email format for {field.get("label", field_name)}'
                            }), 400
                        
                        # Domain validation
                        email_domain = email_value.split('@')[1].lower()
                        if email_domain not in ALLOWED_EMAIL_DOMAINS:
                            allowed_domains_str = ', '.join(ALLOWED_EMAIL_DOMAINS)
                            return jsonify({
                                'error': f'Email domain not allowed. Please use one of: {allowed_domains_str}'
                            }), 400

        # Validate event and registration deadline
        event_id = data.get('event_id')
        event = None
        events = None  # Will hold the events list for saving
        next_id = None  # Will hold the next_id for saving
        
        if event_id is not None:
            try:
                event_id_int = int(event_id)
                events, next_id = load_events_file()
                event = next((e for e in events if e.get('id') == event_id_int), None)
            except (TypeError, ValueError):
                pass
        
        # If no event found by ID, try to find by slug
        if not event:
            events, next_id = load_events_file()
            event = next((e for e in events if slugify(e.get('name', '')) == event_slug), None)
        
        if event:
            # Check registration type
            if event.get('registration_type') not in ['internal']:
                return jsonify({'error': 'Registration is not enabled for this event'}), 400
            
            # Check if registration is allowed (admin toggle)
            if event.get('allow_registration') == False:
                return jsonify({'error': 'Registration is currently closed for this event'}), 400
            
            # Check registration deadline (using IST)
            deadline_info = event.get('registration_deadline')
            if deadline_info and deadline_info.get('date'):
                try:
                    deadline_date = deadline_info['date']
                    # Skip validation for TBA deadlines
                    if deadline_date.upper() != 'TBA':
                        deadline = datetime.strptime(deadline_date, '%Y-%m-%d')
                        # Registration closes after the deadline day (deadline day is last day to register)
                        # Using IST for comparison
                        if deadline.date() < get_ist_now().date():
                            return jsonify({'error': 'Registration deadline has passed'}), 400
                except ValueError:
                    pass
            
            event_slug = slugify(event.get('name', event_slug))

        # Save registration to file
        registrations_dir = os.path.join(PROJECT_ROOT, 'data', 'registrations')
        os.makedirs(registrations_dir, exist_ok=True)
        
        # Get or create registration file for this event
        # Check if event already has a registration file path
        if event and event.get('registration_file'):
            reg_file = os.path.join(PROJECT_ROOT, event['registration_file'])
            logger.debug(f"Using existing registration_file from event: {reg_file}")
        else:
            # Create new registration file with event ID for uniqueness
            event_id = event.get('id', '') if event else ''
            reg_filename = f'{event_slug}_{event_id}_registrations.json'
            reg_file = os.path.join(registrations_dir, reg_filename)
            logger.debug(f"Creating new registration file: {reg_file}")
            
            # Update event with registration file path
            if event and next_id is not None:
                event['registration_file'] = f'data/registrations/{reg_filename}'
                # Save events.json with the updated event using save_events_file
                save_events_file(events, next_id)
                # Reload global EVENTS
                global EVENTS
                EVENTS = events
        
        # NOTE: Duplicate checking is done ONLY in atomic_add_registration to prevent race conditions.
        # Previously there was an early check here, but it caused timing issues where:
        # - Request A passed early check, then saved successfully
        # - Request B passed early check (before A saved), then failed at atomic check
        # - User saw "already registered" error but registration was actually saved by Request A
        # The atomic operation handles all duplicate checking within a single lock.
        
        # Prepare registration data with UUID FIRST (before any file operations)
        registration_uuid = str(uuid.uuid4())
        data['timestamp'] = datetime.now().isoformat()
        data['registration_id'] = registration_uuid
        data['payment_status'] = 'pending' if template_definition and template_definition.get('payment_enabled') else 'not_required'
        data['attendance_status'] = 'not_entered'
        data['entry_time'] = None
        data['marked_by'] = None
        
        # Generate QR code for registration with admin verification URL
        event_name = event.get('name', 'Event') if event else 'Event'
        event_id_param = event.get('id', '') if event else ''
        qr_url = f"{request.host_url}admin/verify-entry?regid={registration_uuid}&email={data.get('submitter_email', '')}&event_id={event_id_param}"
        qr_code_base64 = generate_qr_code(qr_url)
        
        if qr_code_base64:
            data['qr_code'] = qr_code_base64
        
        # Define duplicate check function for atomic operation
        def check_duplicates(registrations, new_reg):
            submitter_email = new_reg.get('submitter_email', '').strip().lower()
            if submitter_email:
                for reg in registrations:
                    existing_email = reg.get('submitter_email', '').strip().lower()
                    if existing_email == submitter_email:
                        return f'Email already registered: {submitter_email}'
            
            # Check other unique fields from template
            if template_definition:
                for field in template_definition.get('fields', []):
                    if field.get('type') == 'email' and field.get('unique', False):
                        field_name = field.get('name')
                        if field_name == 'submitter_email':
                            continue
                        email_value = new_reg.get(field_name, '').strip().lower()
                        if email_value:
                            for reg in registrations:
                                existing_email = reg.get(field_name, '').strip().lower()
                                if existing_email == email_value:
                                    return f'{field.get("label", field_name)} already registered: {email_value}'
            return None  # No duplicates found
        
        # Check if payment is required
        if template_definition and template_definition.get('payment_enabled'):
            payment_amount = template_definition.get('payment_amount', 0)
            if payment_amount > 0:
                # Create Razorpay payment order
                try:
                    # Ensure phone is available
                    customer_phone = data.get('phone', data.get('team_leader_phone', ''))
                    if not customer_phone or not str(customer_phone).strip():
                        return jsonify({
                            'error': 'Phone number is required for payment',
                            'details': 'Please provide a valid phone number'
                        }), 400
                    
                    # BUG FIX: Use registration_uuid instead of data['id'] which isn't set yet
                    order_id = f"ORD_{event_slug}_{registration_uuid[:8]}_{int(time.time())}"
                    payment_order = create_razorpay_order(
                        order_id=order_id,
                        amount=payment_amount,
                        customer_name=data.get('name', data.get('team_leader_name', 'Guest')),
                        customer_email=data.get('email', data.get('team_leader_email', '')),
                        customer_phone=customer_phone,
                        return_url=f"{request.host_url}payment/callback"
                    )
                    
                    # Check if payment order was successful
                    if payment_order and 'order_id' in payment_order and not payment_order.get('error'):
                        # DON'T save registration yet - only save after payment verification
                        # Add payment amount to registration data for verification
                        data['payment_amount'] = payment_amount
                        
                        # Store registration data temporarily in session or return to frontend
                        return jsonify({
                            'success': True,
                            'payment_required': True,
                            'order_id': payment_order['order_id'],
                            'amount': payment_order['amount'],
                            'currency': payment_order['currency'],
                            'key_id': get_razorpay_keys()[0],
                            'registration_data': data,  # Send back to frontend for saving after payment
                            'registration_file': os.path.basename(reg_file)  # Filename to save to
                        }), 200
                    else:
                        # Handle Razorpay error gracefully - return 400, not 500
                        error_msg = 'Failed to create payment order'
                        error_details = 'Payment gateway error. Please contact support.'
                        
                        if payment_order and 'error' in payment_order:
                            if payment_order.get('status_code') == 401:
                                error_details = 'Payment gateway configuration error. Please contact administrator.'
                            elif 'details' in payment_order:
                                error_details = payment_order['details']
                        
                        return jsonify({
                            'error': error_msg,
                            'details': error_details
                        }), 400
                        
                except Exception as e:
                    return jsonify({
                        'error': 'Payment gateway error',
                        'details': 'Unable to process payment. Please try again or contact support.'
                    }), 400
        
        # No payment required - save registration using ATOMIC operation
        # This ensures read-check-write all happens within a single lock
        logger.debug(f"Saving registration to: {reg_file}")
        logger.debug(f"Registration ID being saved: {registration_uuid}")
        
        success, error_msg, _ = atomic_add_registration(reg_file, data, check_duplicates)
        
        if not success:
            return jsonify({
                'error': 'Registration failed',
                'details': error_msg
            }), 400
        
        logger.debug(f"Registration saved successfully with ID: {registration_uuid}")
        
        # Send confirmation email with QR code (use the SAME registration_uuid that was saved)
        email_sent = False
        if qr_code_base64:
            email_sent = send_registration_email(
                email=data.get('submitter_email'),
                registration_id=registration_uuid,  # Use the exact same UUID that was saved
                qr_code_base64=qr_code_base64,
                event_name=event_name,
                registration_data=data
            )
        
        # Return the SAME registration_uuid that was saved to DB and sent in email
        return jsonify({
            'success': True,
            'message': 'Registration submitted successfully!' + (' Confirmation email sent.' if email_sent else ''),
            'registration_id': registration_uuid,  # This must match what's in DB
            'email_sent': email_sent,
            'qr_code': qr_code_base64
        }), 200
        
    except Exception as e:
        logger.exception(f"Registration error: {str(e)}")
        return jsonify({'error': 'Failed to process registration', 'details': str(e)}), 500


@app.route('/payment/verify', methods=['POST'])
def payment_verify():
    """Verify Razorpay payment signature and save registration (Server-side verification)"""
    try:
        data = request.get_json()
        razorpay_payment_id = data.get('razorpay_payment_id')
        razorpay_order_id = data.get('razorpay_order_id')
        razorpay_signature = data.get('razorpay_signature')
        registration_data = data.get('registration_data')
        registration_file = data.get('registration_file')
        
        if not all([razorpay_payment_id, razorpay_order_id, razorpay_signature]):
            return jsonify({'error': 'Missing payment details'}), 400
        
        if not registration_data:
            return jsonify({'error': 'Missing registration data'}), 400
        
        # STEP 1: SERVER-SIDE SIGNATURE VERIFICATION
        # This is critical security step - never trust client-side verification alone
        # hmac and hashlib are imported at module level
        
        _, key_secret = get_razorpay_keys()
        message = f"{razorpay_order_id}|{razorpay_payment_id}"
        expected_signature = hmac.new(
            key_secret.encode(),
            message.encode(),
            hashlib.sha256
        ).hexdigest()
        
        if expected_signature != razorpay_signature:
            # Log failed verification attempt
            logger.warning(f"Payment verification failed for order {razorpay_order_id}")
            return jsonify({'error': 'Invalid payment signature'}), 400
        
        # STEP 2: ADDITIONAL SERVER-SIDE CHECK - Verify payment status with Razorpay API
        # This prevents replay attacks and ensures payment is actually captured
        try:
            verify_url = f"https://api.razorpay.com/v1/payments/{razorpay_payment_id}"
            key_id, key_secret = get_razorpay_keys()
            auth = (key_id, key_secret)
            response = requests.get(verify_url, auth=auth)
            
            if response.status_code == 200:
                payment_details = response.json()
                payment_status = payment_details.get('status')
                
                # Verify payment is actually captured
                if payment_status != 'captured':
                    return jsonify({
                        'error': 'Payment not captured',
                        'status': payment_status
                    }), 400
                
                # Verify order_id matches
                if payment_details.get('order_id') != razorpay_order_id:
                    return jsonify({'error': 'Order ID mismatch'}), 400
                
                # Verify amount matches (prevent tampering)
                expected_amount = int(float(registration_data.get('payment_amount', 0)) * 100)
                actual_amount = payment_details.get('amount', 0)
                
                if expected_amount != actual_amount:
                    logger.warning(f"Amount mismatch - Expected: {expected_amount}, Received: {actual_amount}, Registration Amount: {registration_data.get('payment_amount')}")
                    return jsonify({
                        'error': 'Payment amount mismatch',
                        'details': f'Expected ₹{expected_amount/100}, received ₹{actual_amount/100}'
                    }), 400
            else:
                return jsonify({'error': 'Unable to verify payment with Razorpay'}), 400
                
        except Exception as e:
            logger.error(f"Error verifying payment with Razorpay API: {str(e)}")
            return jsonify({'error': 'Payment verification failed'}), 500
        
        # STEP 3: Payment fully verified on server - NOW save the registration
        registrations_dir = os.path.join(PROJECT_ROOT, 'data', 'registrations')
        
        if registration_file:
            reg_file = os.path.join(registrations_dir, registration_file)
        else:
            return jsonify({'error': 'Missing registration file'}), 400
        
        # Add payment details to registration data
        registration_data['payment_status'] = 'completed'
        registration_data['payment_id'] = razorpay_payment_id
        registration_data['payment_order_id'] = razorpay_order_id
        registration_data['payment_completed_at'] = datetime.now().isoformat()
        registration_data['payment_verified_server_side'] = True
        registration_data['attendance_status'] = 'not_entered'
        registration_data['entry_time'] = None
        registration_data['marked_by'] = None
        
        # Get event name for email
        event_name = 'Event'
        try:
            _, events, _, _ = load_data()
            event_id = registration_data.get('event_id')
            if event_id:
                event = next((e for e in events if e.get('id') == int(event_id)), None)
                if event:
                    event_name = event.get('name', 'Event')
        except:
            pass
        
        # The registration_id should already be in registration_data from the initial form submission
        if 'registration_id' not in registration_data or not registration_data['registration_id']:
            registration_uuid = str(uuid.uuid4())
            registration_data['registration_id'] = registration_uuid
            logger.warning(f"registration_id was missing, generated new one: {registration_uuid}")
        else:
            registration_uuid = registration_data['registration_id']
            logger.debug(f"Using existing registration_id: {registration_uuid}")
        
        # Generate QR code if not present
        if 'qr_code' not in registration_data or not registration_data['qr_code']:
            event_id_param = registration_data.get('event_id', '')
            qr_url = f"{request.host_url}admin/verify-entry?regid={registration_uuid}&email={registration_data.get('submitter_email', '')}&event_id={event_id_param}"
            qr_code_base64 = generate_qr_code(qr_url)
            if qr_code_base64:
                registration_data['qr_code'] = qr_code_base64
        else:
            qr_code_base64 = registration_data['qr_code']
        
        # Define duplicate check function for payment registration
        def check_payment_duplicates(registrations, new_reg):
            # Check for duplicate payment ID
            for reg in registrations:
                if reg.get('payment_id') == razorpay_payment_id:
                    return f'Payment already processed'
                # Also check for duplicate email
                if reg.get('submitter_email', '').lower() == new_reg.get('submitter_email', '').lower():
                    return f'Email already registered'
            return None
        
        # Save registration using ATOMIC operation
        logger.debug(f"Saving payment registration to: {reg_file}")
        logger.debug(f"Registration ID being saved: {registration_uuid}")
        
        success, error_msg, _ = atomic_add_registration(reg_file, registration_data, check_payment_duplicates)
        
        if not success:
            return jsonify({
                'error': error_msg,
                'registration_id': registration_uuid if 'already processed' in (error_msg or '') else None
            }), 400
        
        logger.debug(f"Payment registration saved successfully with ID: {registration_uuid}")
        
        # Send confirmation email with QR code (use the SAME registration_uuid that was saved)
        email_sent = False
        if qr_code_base64:
            email_sent = send_registration_email(
                email=registration_data.get('submitter_email'),
                registration_id=registration_uuid,  # Use the exact same UUID that was saved
                qr_code_base64=qr_code_base64,
                event_name=event_name,
                registration_data=registration_data
            )
        
        return jsonify({
            'success': True,
            'message': 'Payment verified and registration completed!' + (' Confirmation email sent.' if email_sent else ''),
            'registration_id': registration_uuid,
            'email_sent': email_sent,
            'qr_code': qr_code_base64
        }), 200
            
    except Exception as e:
        logger.error(f"Payment verification error: {str(e)}")
        return jsonify({'error': 'Payment verification failed'}), 500


@app.route('/payment/webhook', methods=['POST'])
def payment_webhook():
    """Handle Razorpay webhook for payment notifications (Server-side)"""
    try:
        # Get webhook data - SECURITY: Use environment variable for webhook secret
        webhook_secret = os.environ.get('RAZORPAY_WEBHOOK_SECRET')
        if not webhook_secret:
            logger.error("RAZORPAY_WEBHOOK_SECRET not configured")
            return jsonify({'error': 'Webhook not configured'}), 500
        
        webhook_signature = request.headers.get('X-Razorpay-Signature')
        webhook_body = request.get_data()
        
        # Verify webhook signature (Server-side verification)
        # hmac and hashlib are imported at module level
        expected_signature = hmac.new(
            webhook_secret.encode(),
            webhook_body,
            hashlib.sha256
        ).hexdigest()
        
        if webhook_signature != expected_signature:
            logger.warning("Webhook signature verification failed")
            return jsonify({'error': 'Invalid signature'}), 400
        
        # Process webhook event
        event = request.get_json()
        event_type = event.get('event')
        
        if event_type == 'payment.captured':
            # Payment successful - update registration on server
            payment_entity = event.get('payload', {}).get('payment', {}).get('entity', {})
            order_id = payment_entity.get('order_id')
            payment_id = payment_entity.get('id')
            amount = payment_entity.get('amount')
            
            logger.info(f"Payment captured: {payment_id} for order: {order_id}")
            
            # Find and update registration
            registrations_dir = os.path.join(PROJECT_ROOT, 'data', 'registrations')
            if os.path.exists(registrations_dir):
                for filename in os.listdir(registrations_dir):
                    if filename.endswith('_registrations.json'):
                        filepath = os.path.join(registrations_dir, filename)
                        registrations = safe_json_read(filepath)
                        
                        for reg in registrations:
                            if reg.get('payment_order_id') == order_id:
                                reg['payment_status'] = 'completed'
                                reg['payment_id'] = payment_id
                                reg['payment_completed_at'] = datetime.now().isoformat()
                                reg['webhook_verified'] = True
                                
                                safe_json_write(filepath, registrations)
                                
                                return jsonify({'status': 'ok'}), 200
        
        elif event_type == 'payment.failed':
            # Payment failed - update status
            payment_entity = event.get('payload', {}).get('payment', {}).get('entity', {})
            order_id = payment_entity.get('order_id')
            
            logger.warning(f"Payment failed for order: {order_id}")
            
            # Update registration status
            registrations_dir = os.path.join(PROJECT_ROOT, 'data', 'registrations')
            if os.path.exists(registrations_dir):
                for filename in os.listdir(registrations_dir):
                    if filename.endswith('_registrations.json'):
                        filepath = os.path.join(registrations_dir, filename)
                        registrations = safe_json_read(filepath)
                        
                        for reg in registrations:
                            if reg.get('payment_order_id') == order_id:
                                reg['payment_status'] = 'failed'
                                reg['payment_failed_at'] = datetime.now().isoformat()
                                
                                safe_json_write(filepath, registrations)
                                
                                return jsonify({'status': 'ok'}), 200
        
        return jsonify({'status': 'ok'}), 200
        
    except Exception as e:
        logger.error(f"Webhook processing error: {str(e)}")
        return jsonify({'error': 'Webhook processing failed'}), 500


@app.route('/payment/status/<order_id>', methods=['GET'])
def payment_status(order_id):
    """Check payment status from Razorpay (Server-side check)"""
    try:
        # Verify with Razorpay API
        verify_url = f"https://api.razorpay.com/v1/orders/{order_id}"
        key_id, key_secret = get_razorpay_keys()
        auth = (key_id, key_secret)
        response = requests.get(verify_url, auth=auth)
        
        if response.status_code == 200:
            order_details = response.json()
            return jsonify({
                'success': True,
                'order_id': order_details.get('id'),
                'status': order_details.get('status'),
                'amount': order_details.get('amount'),
                'amount_paid': order_details.get('amount_paid'),
                'attempts': order_details.get('attempts')
            }), 200
        else:
            return jsonify({'error': 'Unable to fetch order status'}), 400
            
    except Exception as e:
        logger.error(f"Status check error: {str(e)}")
        return jsonify({'error': 'Status check failed'}), 500


@app.route('/payment/callback')
def payment_callback():
    """Handle Razorpay payment callback/redirect"""
    payment_id = request.args.get('razorpay_payment_id')
    order_id = request.args.get('razorpay_order_id')
    signature = request.args.get('razorpay_signature')
    
    if not all([payment_id, order_id, signature]):
        flash('Invalid payment callback', 'error')
        return redirect(url_for('events'))
    
    # Verify signature on server
    try:
        # hmac and hashlib are imported at module level
        _, key_secret = get_razorpay_keys()
        message = f"{order_id}|{payment_id}"
        expected_signature = hmac.new(
            key_secret.encode(),
            message.encode(),
            hashlib.sha256
        ).hexdigest()
        
        if signature == expected_signature:
            flash('Payment successful! Your registration is confirmed.', 'success')
        else:
            flash('Payment verification failed. Please contact support.', 'error')
            
    except Exception as e:
        flash('Error processing payment. Please contact support.', 'error')
    
    return redirect(url_for('events'))


@app.route('/members')
def members():
    """Members page showing team members"""
    global CLUB_INFO, EVENTS, MEMBERS, GALLERY
    CLUB_INFO, EVENTS, MEMBERS, GALLERY = load_data()
    return render_template('members.html', 
                         members=MEMBERS,
                         club_info=CLUB_INFO,
                         contact=CLUB_INFO)

@app.route('/gallery')
def gallery():
    """Life @ AICC gallery page"""
    global CLUB_INFO, EVENTS, MEMBERS, GALLERY
    CLUB_INFO, EVENTS, MEMBERS, GALLERY = load_data()
    return render_template('gallery.html', 
                         gallery=GALLERY,
                         club_info=CLUB_INFO,
                         contact=CLUB_INFO)

@app.route('/api/events')
def api_events():
    """API endpoint to get events data"""
    global CLUB_INFO, EVENTS, MEMBERS, GALLERY
    CLUB_INFO, EVENTS, MEMBERS, GALLERY = load_data()
    return jsonify(EVENTS)


@app.route('/api/members')
def api_members():
    """API endpoint to get members data"""
    global CLUB_INFO, EVENTS, MEMBERS, GALLERY
    CLUB_INFO, EVENTS, MEMBERS, GALLERY = load_data()
    return jsonify(MEMBERS)

@app.route('/api/data')
def api_data():
    """Bulk API endpoint: returns ALL public data in a single response.
    Used by the React frontend to minimize API calls (CPU-saving for PythonAnywhere free tier).
    """
    global CLUB_INFO, EVENTS, MEMBERS, GALLERY
    CLUB_INFO, EVENTS, MEMBERS, GALLERY = load_data()
    
    # Strip sensitive fields from club info
    sensitive_keys = {'api_config', 'email_config', 'admin_password'}
    safe_club_info = {k: v for k, v in CLUB_INFO.items() if k not in sensitive_keys}
    
    # Load form templates (active only)
    templates_file = os.path.join(PROJECT_ROOT, 'data', 'form_templates.json')
    form_templates = []
    if os.path.exists(templates_file):
        try:
            with open(templates_file, 'r') as f:
                all_templates = json.load(f)
            form_templates = [t for t in all_templates if t.get('active')]
        except Exception:
            form_templates = []
    
    return jsonify({
        'club': safe_club_info,
        'events': EVENTS,
        'members': MEMBERS,
        'gallery': GALLERY,
        'form_templates': form_templates
    })

@app.route('/api/attendance/check', methods=['POST'])
def api_attendance_check():
    """JSON-only attendance check API for the React frontend."""
    global CLUB_INFO, EVENTS, MEMBERS, GALLERY
    CLUB_INFO, EVENTS, MEMBERS, GALLERY = load_data()
    
    data = request.get_json(silent=True) or {}
    email = data.get('email', '').strip().lower()
    reg_id = data.get('registration_id', '').strip()
    event_id = data.get('event_id')
    
    if not email or not reg_id or not event_id:
        return jsonify({'error': 'Please provide email, registration_id, and event_id.'}), 400
    
    try:
        event_id = int(event_id)
        event = next((e for e in EVENTS if e.get('id') == event_id), None)
        
        if not event:
            return jsonify({'error': 'Event not found.'}), 404
        
        # Load registrations for this event
        registrations = []
        if event.get('registration_file'):
            reg_file_path = os.path.join(PROJECT_ROOT, event['registration_file'])
            if os.path.exists(reg_file_path):
                with open(reg_file_path, 'r') as f:
                    registrations = json.load(f)
        else:
            event_slug = slugify(event.get('name', ''))
            reg_file_path = os.path.join(PROJECT_ROOT, 'data', 'registrations', f'{event_slug}_registrations.json')
            if os.path.exists(reg_file_path):
                with open(reg_file_path, 'r') as f:
                    registrations = json.load(f)
        
        # Find the registration
        registration = None
        for reg in registrations:
            if reg.get('registration_id') == reg_id and reg.get('submitter_email', '').lower() == email:
                registration = reg
                break
        
        if not registration:
            return jsonify({'error': 'Registration not found. Please check your email and registration ID.'}), 404
        
        # Build participant name
        name = registration.get('name', registration.get('submitter_email', 'Participant'))
        if registration.get('participants') and len(registration['participants']) > 0:
            name = registration['participants'][0].get('name', name)
        
        return jsonify({
            'success': True,
            'event': {
                'id': event.get('id'),
                'name': event.get('name'),
                'date': event.get('date'),
            },
            'registration': {
                'registration_id': registration.get('registration_id'),
                'submitter_email': registration.get('submitter_email'),
                'participants': registration.get('participants', []),
                'timestamp': registration.get('timestamp'),
                'custom_field_values': registration.get('custom_field_values', {}),
                'num_participants': registration.get('num_participants', 1),
                'participant_attendance': registration.get('participant_attendance', []),
            },
            'name': name,
            'attendance_status': registration.get('attendance_status', 'not_entered'),
            'entry_time': registration.get('entry_time'),
            'attendance_comment': registration.get('attendance_comment', ''),
            'marked_by': registration.get('marked_by', ''),
        })
        
    except (ValueError, TypeError):
        return jsonify({'error': 'Invalid event ID.'}), 400

# ========================================
# Attendance Check Routes (Public)
# ========================================

@app.route('/attendance/check', methods=['GET', 'POST'])
def attendance_check():
    """Public page for attendees to check their attendance status
    
    Supports GET parameters for shareable links:
    - event_id: The event ID
    - email: The attendee's email
    - registration_id or rid: The registration UUID
    
    Example: /attendance/check?event_id=1&email=test@example.com&rid=66446360-a634-4179-904f-c77100275e76
    """
    global CLUB_INFO, EVENTS, MEMBERS, GALLERY
    CLUB_INFO, EVENTS, MEMBERS, GALLERY = load_data()
    
    attendance_info = None
    error_message = None
    
    # Check for GET parameters (shareable link) or POST form submission
    if request.method == 'POST':
        email = request.form.get('email', '').strip().lower()
        reg_id = request.form.get('registration_id', '').strip()
        event_id = request.form.get('event_id')
    else:
        # GET request - check for query parameters
        email = request.args.get('email', '').strip().lower()
        reg_id = request.args.get('registration_id', request.args.get('rid', '')).strip()
        event_id = request.args.get('event_id')
    
    # Process if we have the required parameters (either from GET or POST)
    if email and reg_id and event_id:
        try:
            event_id = int(event_id)
            event = next((e for e in EVENTS if e.get('id') == event_id), None)
            
            if not event:
                error_message = 'Event not found.'
            else:
                # Load registrations for this event
                registrations = []
                reg_file_path = None
                
                if event.get('registration_file'):
                    reg_file_path = os.path.join(PROJECT_ROOT, event['registration_file'])
                    if os.path.exists(reg_file_path):
                        with open(reg_file_path, 'r') as f:
                            registrations = json.load(f)
                else:
                    event_slug = slugify(event.get('name', ''))
                    reg_file_path = os.path.join(PROJECT_ROOT, 'data', 'registrations', f'{event_slug}_registrations.json')
                    if os.path.exists(reg_file_path):
                        with open(reg_file_path, 'r') as f:
                            registrations = json.load(f)
                
                # Find the registration
                registration = None
                for reg in registrations:
                    if reg.get('registration_id') == reg_id and reg.get('submitter_email', '').lower() == email:
                        registration = reg
                        break
                
                if not registration:
                    error_message = 'Registration not found. Please check your email and registration ID.'
                else:
                    attendance_info = {
                        'event': event,
                        'registration': registration,
                        'name': registration.get('name', registration.get('submitter_email', 'Participant')),
                        'attendance_status': registration.get('attendance_status', 'not_entered'),
                        'entry_time': registration.get('entry_time'),
                        'attendance_comment': registration.get('attendance_comment', ''),
                        'marked_by': registration.get('marked_by', ''),
                        'total_registrations': len(registrations)
                    }
        except (ValueError, TypeError):
            error_message = 'Invalid event selection.'
    elif request.method == 'POST':
        # Only show error for POST with missing fields
        error_message = 'Please provide email, registration ID, and select an event.'
    
    # Generate shareable link if we have attendance info
    shareable_link = None
    shareable_qr_code = None
    if attendance_info:
        shareable_link = url_for('attendance_check', 
                                  event_id=attendance_info['event']['id'],
                                  email=attendance_info['registration']['submitter_email'],
                                  rid=attendance_info['registration']['registration_id'],
                                  _external=True)
        
        # Generate QR code for the shareable link
        try:
            qr = qrcode.QRCode(version=1, box_size=10, border=4)
            qr.add_data(shareable_link)
            qr.make(fit=True)
            qr_img = qr.make_image(fill_color="black", back_color="white")
            
            # Convert to base64
            buffered = io.BytesIO()
            qr_img.save(buffered, format="PNG")
            shareable_qr_code = base64.b64encode(buffered.getvalue()).decode('utf-8')
        except Exception as e:
            logging.error(f"Error generating shareable QR code: {e}")
            shareable_qr_code = None
    
    return render_template('attendance_check.html',
                         club_info=CLUB_INFO,
                         contact=CLUB_INFO,
                         events=EVENTS,
                         attendance_info=attendance_info,
                         error_message=error_message,
                         shareable_link=shareable_link,
                         shareable_qr_code=shareable_qr_code)

# ========================================
# Admin API Routes (Token-based auth for React frontend)
# ========================================

import secrets

# In-memory admin tokens (simple approach - token -> expiry timestamp)
_admin_tokens = {}

def _cleanup_tokens():
    """Remove expired tokens"""
    now = time.time()
    expired = [t for t, exp in _admin_tokens.items() if exp < now]
    for t in expired:
        del _admin_tokens[t]

def _generate_admin_token():
    """Generate a secure admin token valid for 24 hours"""
    _cleanup_tokens()
    token = secrets.token_urlsafe(32)
    _admin_tokens[token] = time.time() + 86400  # 24h
    return token

def _verify_admin_token(token):
    """Verify an admin token is valid"""
    _cleanup_tokens()
    return token in _admin_tokens

def api_admin_required(f):
    """Decorator for API routes requiring admin token"""
    @wraps(f)
    def decorated(*args, **kwargs):
        auth = request.headers.get('Authorization', '')
        token = auth.replace('Bearer ', '') if auth.startswith('Bearer ') else ''
        if not token or not _verify_admin_token(token):
            return jsonify({'error': 'Unauthorized'}), 401
        return f(*args, **kwargs)
    return decorated

@app.route('/api/admin/login', methods=['POST'])
def api_admin_login():
    """Admin login - returns token"""
    data = request.get_json(silent=True) or {}
    username = data.get('username', '')
    password = data.get('password', '')
    if username == ADMIN_USERNAME and password == ADMIN_PASSWORD:
        token = _generate_admin_token()
        return jsonify({'success': True, 'token': token})
    return jsonify({'error': 'Invalid credentials'}), 401

@app.route('/api/admin/verify', methods=['GET'])
def api_admin_verify():
    """Verify admin token is still valid"""
    auth = request.headers.get('Authorization', '')
    token = auth.replace('Bearer ', '') if auth.startswith('Bearer ') else ''
    if token and _verify_admin_token(token):
        return jsonify({'valid': True})
    return jsonify({'valid': False}), 401

@app.route('/api/admin/dashboard', methods=['GET'])
@api_admin_required
def api_admin_dashboard():
    """Get dashboard stats"""
    global CLUB_INFO, EVENTS, MEMBERS, GALLERY
    CLUB_INFO, EVENTS, MEMBERS, GALLERY = load_data()
    return jsonify({
        'events_count': len(EVENTS),
        'members_count': len(MEMBERS),
        'gallery_count': len(GALLERY),
    })

@app.route('/api/admin/club-info', methods=['GET', 'PUT'])
@api_admin_required
def api_admin_club_info():
    """Get or update club information"""
    global CLUB_INFO, EVENTS, MEMBERS, GALLERY
    CLUB_INFO, EVENTS, MEMBERS, GALLERY = load_data()
    
    if request.method == 'GET':
        return jsonify(CLUB_INFO)
    
    # PUT - update
    data = request.get_json(silent=True) or {}
    # Merge with existing, preserving keys not in request
    for key in data:
        CLUB_INFO[key] = data[key]
    
    with open(os.path.join(PROJECT_ROOT, 'data/club_info.json'), 'w') as f:
        json.dump(CLUB_INFO, f, indent=4)
    CLUB_INFO, EVENTS, MEMBERS, GALLERY = load_data()
    return jsonify({'success': True})

@app.route('/api/admin/events', methods=['GET'])
@api_admin_required
def api_admin_events():
    """Get all events for admin"""
    global CLUB_INFO, EVENTS, MEMBERS, GALLERY
    CLUB_INFO, EVENTS, MEMBERS, GALLERY = load_data()
    return jsonify(EVENTS)

@app.route('/api/admin/events', methods=['POST'])
@api_admin_required
def api_admin_create_event():
    """Create a new event via API"""
    global CLUB_INFO, EVENTS, MEMBERS, GALLERY
    data = request.get_json(silent=True) or {}
    
    events, next_id = load_events_file()
    
    new_event = {
        'id': next_id,
        'name': data.get('name', ''),
        'date': data.get('date', ''),
        'time': data.get('time', ''),
        'location': data.get('location', ''),
        'description': data.get('description', ''),
        'how': data.get('how', ''),
        'status': data.get('status', 'upcoming'),
        'image': data.get('image', ''),
        'rules': data.get('rules', []),
        'coordinators': data.get('coordinators', []),
        'registration_type': data.get('registration_type', 'none'),
        'register_link': data.get('register_link', '#'),
        'template_id': data.get('template_id'),
        'show_in_events': data.get('show_in_events', True),
    }
    
    # Handle registration deadline
    if data.get('registration_deadline'):
        new_event['registration_deadline'] = data['registration_deadline']
    
    # Create registration file for internal registration
    if new_event['registration_type'] == 'internal' and new_event.get('template_id'):
        event_slug = re.sub(r'[^a-z0-9]+', '_', new_event['name'].lower()).strip('_')
        reg_filename = f"{event_slug}_{new_event['id']}_registrations.json"
        reg_file_path = os.path.join(PROJECT_ROOT, 'data', 'registrations', reg_filename)
        os.makedirs(os.path.dirname(reg_file_path), exist_ok=True)
        if not os.path.exists(reg_file_path):
            with open(reg_file_path, 'w') as f:
                json.dump([], f)
        new_event['registration_file'] = f'data/registrations/{reg_filename}'
    
    events.append(new_event)
    save_events_file(events, next_id + 1)
    CLUB_INFO, EVENTS, MEMBERS, GALLERY = load_data()
    return jsonify({'success': True, 'event': new_event})

@app.route('/api/admin/events/<int:event_id>', methods=['PUT'])
@api_admin_required
def api_admin_update_event(event_id):
    """Update an event via API"""
    global CLUB_INFO, EVENTS, MEMBERS, GALLERY
    data = request.get_json(silent=True) or {}
    
    events, next_id = load_events_file()
    event = next((e for e in events if e.get('id') == event_id), None)
    if not event:
        return jsonify({'error': 'Event not found'}), 404
    
    # Update fields that are provided
    for key in ['name', 'date', 'time', 'location', 'description', 'how', 'status',
                'image', 'rules', 'coordinators', 'registration_type', 'register_link',
                'template_id', 'registration_deadline', 'allow_registration', 'show_in_events']:
        if key in data:
            event[key] = data[key]
    
    # Create registration file if switching to internal
    if event.get('registration_type') == 'internal' and event.get('template_id') and not event.get('registration_file'):
        event_slug = re.sub(r'[^a-z0-9]+', '_', event['name'].lower()).strip('_')
        reg_filename = f"{event_slug}_{event['id']}_registrations.json"
        reg_file_path = os.path.join(PROJECT_ROOT, 'data', 'registrations', reg_filename)
        os.makedirs(os.path.dirname(reg_file_path), exist_ok=True)
        if not os.path.exists(reg_file_path):
            with open(reg_file_path, 'w') as f:
                json.dump([], f)
        event['registration_file'] = f'data/registrations/{reg_filename}'
    
    save_events_file(events, next_id)
    CLUB_INFO, EVENTS, MEMBERS, GALLERY = load_data()
    return jsonify({'success': True, 'event': event})

@app.route('/api/admin/events/<int:event_id>', methods=['DELETE'])
@api_admin_required
def api_admin_delete_event(event_id):
    """Archive an event (mark as completed)"""
    global CLUB_INFO, EVENTS, MEMBERS, GALLERY
    events, next_id = load_events_file()
    event = next((e for e in events if e.get('id') == event_id), None)
    if event:
        event['status'] = 'completed'
        event['registration_type'] = 'none'
        event['allow_registration'] = False
    save_events_file(events, next_id)
    CLUB_INFO, EVENTS, MEMBERS, GALLERY = load_data()
    return jsonify({'success': True})

@app.route('/api/admin/events/<int:event_id>/registrations', methods=['GET'])
@api_admin_required
def api_admin_event_registrations(event_id):
    """Get registrations for an event"""
    global CLUB_INFO, EVENTS, MEMBERS, GALLERY
    CLUB_INFO, EVENTS, MEMBERS, GALLERY = load_data()
    
    event = next((e for e in EVENTS if e.get('id') == event_id), None)
    if not event:
        return jsonify({'error': 'Event not found'}), 404
    
    registrations = []
    if event.get('registration_file'):
        reg_file = os.path.join(PROJECT_ROOT, event['registration_file'])
        if os.path.exists(reg_file):
            with open(reg_file, 'r') as f:
                registrations = json.load(f)
    
    # Load form template
    template = None
    if event.get('template_id'):
        templates_file = os.path.join(PROJECT_ROOT, 'data', 'form_templates.json')
        try:
            with open(templates_file, 'r') as f:
                templates = json.load(f)
            template = next((t for t in templates if t.get('id') == event.get('template_id')), None)
        except:
            pass
    
    return jsonify({
        'event': event,
        'registrations': registrations,
        'form_template': template,
    })

@app.route('/api/admin/events/<int:event_id>/toggle-registration', methods=['POST'])
@api_admin_required
def api_admin_toggle_registration(event_id):
    """Toggle registration for an event"""
    global CLUB_INFO, EVENTS, MEMBERS, GALLERY
    events, next_id = load_events_file()
    event = next((e for e in events if e.get('id') == event_id), None)
    if not event:
        return jsonify({'error': 'Event not found'}), 404
    event['allow_registration'] = not event.get('allow_registration', True)
    save_events_file(events, next_id)
    CLUB_INFO, EVENTS, MEMBERS, GALLERY = load_data()
    return jsonify({'success': True, 'allow_registration': event['allow_registration']})

@app.route('/api/admin/members', methods=['GET'])
@api_admin_required
def api_admin_members():
    """Get all members"""
    global CLUB_INFO, EVENTS, MEMBERS, GALLERY
    CLUB_INFO, EVENTS, MEMBERS, GALLERY = load_data()
    return jsonify({'members': MEMBERS, 'club_info': {
        'member_roles': CLUB_INFO.get('member_roles', []),
        'member_years': CLUB_INFO.get('member_years', []),
    }})

@app.route('/api/admin/members', methods=['POST'])
@api_admin_required
def api_admin_create_member():
    """Add a new member"""
    global CLUB_INFO, EVENTS, MEMBERS, GALLERY
    data = request.get_json(silent=True) or {}
    
    with open(os.path.join(PROJECT_ROOT, 'data/members.json'), 'r') as f:
        members = json.load(f)
    
    members.append({
        'name': data.get('name', ''),
        'role': data.get('role', ''),
        'year': data.get('year', ''),
        'domain': data.get('domain', ''),
        'image': data.get('image', '/static/img/members/default.webp'),
        'linkedin': data.get('linkedin', ''),
        'github': data.get('github', ''),
    })
    
    role_hierarchy = CLUB_INFO.get('member_roles', [])
    year_hierarchy = CLUB_INFO.get('member_years', [])
    if role_hierarchy:
        members = sort_members_by_role(members, role_hierarchy, year_hierarchy)
    
    with open(os.path.join(PROJECT_ROOT, 'data/members.json'), 'w') as f:
        json.dump(members, f, indent=4)
    CLUB_INFO, EVENTS, MEMBERS, GALLERY = load_data()
    return jsonify({'success': True})

@app.route('/api/admin/members/<int:idx>', methods=['PUT'])
@api_admin_required
def api_admin_update_member(idx):
    """Update a member"""
    global CLUB_INFO, EVENTS, MEMBERS, GALLERY
    data = request.get_json(silent=True) or {}
    
    with open(os.path.join(PROJECT_ROOT, 'data/members.json'), 'r') as f:
        members = json.load(f)
    if idx >= len(members):
        return jsonify({'error': 'Member not found'}), 404
    
    for key in ['name', 'role', 'year', 'domain', 'image', 'linkedin', 'github']:
        if key in data:
            members[idx][key] = data[key]
    
    role_hierarchy = CLUB_INFO.get('member_roles', [])
    year_hierarchy = CLUB_INFO.get('member_years', [])
    if role_hierarchy:
        members = sort_members_by_role(members, role_hierarchy, year_hierarchy)
    
    with open(os.path.join(PROJECT_ROOT, 'data/members.json'), 'w') as f:
        json.dump(members, f, indent=4)
    CLUB_INFO, EVENTS, MEMBERS, GALLERY = load_data()
    return jsonify({'success': True})

@app.route('/api/admin/members/<int:idx>', methods=['DELETE'])
@api_admin_required
def api_admin_delete_member(idx):
    """Delete a member"""
    global CLUB_INFO, EVENTS, MEMBERS, GALLERY
    
    with open(os.path.join(PROJECT_ROOT, 'data/members.json'), 'r') as f:
        members = json.load(f)
    if idx < len(members):
        member = members[idx]
        delete_old_image(member.get('image', ''))
        members.pop(idx)
        with open(os.path.join(PROJECT_ROOT, 'data/members.json'), 'w') as f:
            json.dump(members, f, indent=4)
    CLUB_INFO, EVENTS, MEMBERS, GALLERY = load_data()
    return jsonify({'success': True})

@app.route('/api/admin/gallery', methods=['GET'])
@api_admin_required
def api_admin_gallery():
    """Get all gallery images"""
    global CLUB_INFO, EVENTS, MEMBERS, GALLERY
    CLUB_INFO, EVENTS, MEMBERS, GALLERY = load_data()
    return jsonify(GALLERY)

@app.route('/api/admin/gallery', methods=['POST'])
@api_admin_required
def api_admin_create_gallery():
    """Add a gallery image"""
    global CLUB_INFO, EVENTS, MEMBERS, GALLERY
    data = request.get_json(silent=True) or {}
    
    with open(os.path.join(PROJECT_ROOT, 'data/gallery.json'), 'r') as f:
        gallery = json.load(f)
    
    gallery.append({
        'url': data.get('url', ''),
        'image': data.get('image', data.get('url', '')),
        'title': data.get('title', ''),
        'category': data.get('category', 'events'),
        'description': data.get('description', ''),
    })
    
    with open(os.path.join(PROJECT_ROOT, 'data/gallery.json'), 'w') as f:
        json.dump(gallery, f, indent=4)
    CLUB_INFO, EVENTS, MEMBERS, GALLERY = load_data()
    return jsonify({'success': True})

@app.route('/api/admin/gallery/<int:idx>', methods=['PUT'])
@api_admin_required
def api_admin_update_gallery(idx):
    """Update a gallery image"""
    global CLUB_INFO, EVENTS, MEMBERS, GALLERY
    data = request.get_json(silent=True) or {}
    
    with open(os.path.join(PROJECT_ROOT, 'data/gallery.json'), 'r') as f:
        gallery = json.load(f)
    if idx >= len(gallery):
        return jsonify({'error': 'Image not found'}), 404
    
    for key in ['title', 'category', 'description', 'url', 'image']:
        if key in data:
            gallery[idx][key] = data[key]
    
    with open(os.path.join(PROJECT_ROOT, 'data/gallery.json'), 'w') as f:
        json.dump(gallery, f, indent=4)
    CLUB_INFO, EVENTS, MEMBERS, GALLERY = load_data()
    return jsonify({'success': True})

@app.route('/api/admin/gallery/<int:idx>', methods=['DELETE'])
@api_admin_required
def api_admin_delete_gallery(idx):
    """Delete a gallery image"""
    global CLUB_INFO, EVENTS, MEMBERS, GALLERY
    
    with open(os.path.join(PROJECT_ROOT, 'data/gallery.json'), 'r') as f:
        gallery = json.load(f)
    if idx < len(gallery):
        image = gallery[idx]
        delete_old_image(image.get('url') or image.get('image', ''))
        gallery.pop(idx)
        with open(os.path.join(PROJECT_ROOT, 'data/gallery.json'), 'w') as f:
            json.dump(gallery, f, indent=4)
    CLUB_INFO, EVENTS, MEMBERS, GALLERY = load_data()
    return jsonify({'success': True})

@app.route('/api/admin/contact', methods=['GET', 'PUT'])
@api_admin_required
def api_admin_contact():
    """Get or update contact information"""
    global CLUB_INFO, EVENTS, MEMBERS, GALLERY
    CLUB_INFO, EVENTS, MEMBERS, GALLERY = load_data()
    
    if request.method == 'GET':
        return jsonify({
            'email': CLUB_INFO.get('email', ''),
            'linkedin': CLUB_INFO.get('linkedin', ''),
            'instagram': CLUB_INFO.get('instagram', ''),
            'faculty_coordinators': CLUB_INFO.get('faculty_coordinators', []),
            'secretaries': CLUB_INFO.get('secretaries', []),
        })
    
    data = request.get_json(silent=True) or {}
    for key in ['email', 'linkedin', 'instagram', 'faculty_coordinators', 'secretaries']:
        if key in data:
            CLUB_INFO[key] = data[key]
    
    with open(os.path.join(PROJECT_ROOT, 'data/club_info.json'), 'w') as f:
        json.dump(CLUB_INFO, f, indent=4)
    CLUB_INFO, EVENTS, MEMBERS, GALLERY = load_data()
    return jsonify({'success': True})

@app.route('/api/admin/form-templates', methods=['GET'])
@api_admin_required
def api_admin_form_templates():
    """Get all form templates"""
    templates_file = os.path.join(PROJECT_ROOT, 'data', 'form_templates.json')
    templates = []
    if os.path.exists(templates_file):
        with open(templates_file, 'r') as f:
            templates = json.load(f)
    return jsonify(templates)

@app.route('/api/admin/form-templates', methods=['POST'])
@api_admin_required
def api_admin_create_form_template():
    """Create a form template"""
    data = request.get_json(silent=True) or {}
    templates_file = os.path.join(PROJECT_ROOT, 'data', 'form_templates.json')
    templates = []
    if os.path.exists(templates_file):
        with open(templates_file, 'r') as f:
            templates = json.load(f)
    
    max_id = max([t.get('id', 0) for t in templates], default=0)
    data['id'] = max_id + 1
    templates.append(data)
    
    with open(templates_file, 'w') as f:
        json.dump(templates, f, indent=4)
    return jsonify({'success': True, 'id': data['id']})

@app.route('/api/admin/form-templates/<int:form_id>', methods=['PUT'])
@api_admin_required
def api_admin_update_form_template(form_id):
    """Update a form template"""
    data = request.get_json(silent=True) or {}
    templates_file = os.path.join(PROJECT_ROOT, 'data', 'form_templates.json')
    
    with open(templates_file, 'r') as f:
        templates = json.load(f)
    
    template = next((t for t in templates if t.get('id') == form_id), None)
    if not template:
        return jsonify({'error': 'Template not found'}), 404
    
    for key in data:
        if key != 'id':
            template[key] = data[key]
    
    with open(templates_file, 'w') as f:
        json.dump(templates, f, indent=4)
    return jsonify({'success': True})

@app.route('/api/admin/form-templates/<int:form_id>', methods=['DELETE'])
@api_admin_required
def api_admin_delete_form_template(form_id):
    """Delete a form template"""
    templates_file = os.path.join(PROJECT_ROOT, 'data', 'form_templates.json')
    
    with open(templates_file, 'r') as f:
        templates = json.load(f)
    
    templates = [t for t in templates if t.get('id') != form_id]
    
    with open(templates_file, 'w') as f:
        json.dump(templates, f, indent=4)
    return jsonify({'success': True})

@app.route('/api/admin/form-templates/<int:form_id>/toggle', methods=['POST'])
@api_admin_required
def api_admin_toggle_form_template(form_id):
    """Toggle form template active status"""
    templates_file = os.path.join(PROJECT_ROOT, 'data', 'form_templates.json')
    
    with open(templates_file, 'r') as f:
        templates = json.load(f)
    
    template = next((t for t in templates if t.get('id') == form_id), None)
    if not template:
        return jsonify({'error': 'Template not found'}), 404
    
    template['active'] = not template.get('active', True)
    
    with open(templates_file, 'w') as f:
        json.dump(templates, f, indent=4)
    return jsonify({'success': True, 'active': template['active']})

@app.route('/api/admin/mark-entry', methods=['POST'])
@api_admin_required
def api_admin_mark_entry():
    """Mark attendee entry via API"""
    data = request.get_json(silent=True) or {}
    regid = data.get('regid')
    email = data.get('email')
    event_id = data.get('event_id')
    attendance_type = data.get('attendance_type', 'full')
    attendance_comment = data.get('attendance_comment', '')
    participant_attendance = data.get('participant_attendance')
    
    if not regid or not email or not event_id:
        return jsonify({'error': 'Missing parameters'}), 400
    
    try:
        event_id = int(event_id)
    except (ValueError, TypeError):
        return jsonify({'error': 'Invalid event ID'}), 400
    
    global CLUB_INFO, EVENTS, MEMBERS, GALLERY
    CLUB_INFO, EVENTS, MEMBERS, GALLERY = load_data()
    
    event = next((e for e in EVENTS if e.get('id') == event_id), None)
    if not event:
        return jsonify({'error': 'Event not found'}), 404
    
    reg_file_path = None
    if event.get('registration_file'):
        reg_file_path = os.path.join(PROJECT_ROOT, event['registration_file'])
    else:
        event_slug = slugify(event.get('name', ''))
        reg_file_path = os.path.join(PROJECT_ROOT, 'data', 'registrations', f'{event_slug}_registrations.json')
    
    registrations = safe_json_read(reg_file_path)
    
    updated = False
    for reg in registrations:
        if reg.get('registration_id') == regid and reg.get('submitter_email', '').lower() == email.lower():
            if attendance_type == 'participants' and participant_attendance:
                reg['participant_attendance'] = participant_attendance
                total = len(participant_attendance)
                present = sum(1 for p in participant_attendance if p)
                if present == total:
                    reg['attendance_status'] = 'entered'
                elif present > 0:
                    reg['attendance_status'] = 'partially_present'
                else:
                    reg['attendance_status'] = 'not_entered'
                reg['attendance_comment'] = f'{present}/{total} participants present'
            else:
                reg['attendance_status'] = 'partially_present' if attendance_type == 'partial' else 'entered'
                reg['attendance_comment'] = attendance_comment
            
            reg['entry_time'] = datetime.now().isoformat()
            reg['marked_by'] = 'admin'
            updated = True
            break
    
    if not updated:
        return jsonify({'error': 'Registration not found'}), 404
    
    try:
        safe_json_write(reg_file_path, registrations)
        return jsonify({'success': True})
    except Exception:
        return jsonify({'error': 'Failed to save'}), 500

@app.route('/api/admin/upload', methods=['POST'])
@api_admin_required
def api_admin_upload():
    """Handle file uploads via API"""
    if 'file' not in request.files:
        return jsonify({'error': 'No file provided'}), 400
    file = request.files['file']
    if file.filename == '':
        return jsonify({'error': 'No file selected'}), 400
    if file and allowed_file(file.filename):
        filename = secure_filename(file.filename)
        timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
        filename = f"{timestamp}_{filename}"
        os.makedirs(app.config['UPLOAD_FOLDER'], exist_ok=True)
        filepath = os.path.join(app.config['UPLOAD_FOLDER'], filename)
        file.save(filepath)
        return jsonify({'url': f"/static/uploads/{filename}"})
    return jsonify({'error': 'Invalid file type'}), 400

# ========================================
# Admin Routes (Session-based for Jinja templates)
# ========================================

@app.route('/admin/login', methods=['GET', 'POST'])
def admin_login():
    """Admin login page"""
    if request.method == 'POST':
        username = request.form.get('username')
        password = request.form.get('password')
        
        if username == ADMIN_USERNAME and password == ADMIN_PASSWORD:
            session['admin_logged_in'] = True
            flash('Successfully logged in!', 'success')
            # Redirect to 'next' URL if provided, otherwise to dashboard
            next_url = request.args.get('next') or request.form.get('next')
            if next_url and next_url.startswith('/'):
                return redirect(next_url)
            return redirect(url_for('admin_dashboard'))
        else:
            flash('Invalid credentials. Please try again.', 'error')
    
    return render_template('admin/login.html', next=request.args.get('next', ''))

@app.route('/admin/logout')
def admin_logout():
    """Admin logout"""
    session.pop('admin_logged_in', None)
    flash('Successfully logged out.', 'success')
    return redirect(url_for('admin_login'))

def admin_required(f):
    """Decorator to require admin login"""
    @wraps(f)
    def decorated_function(*args, **kwargs):
        if not session.get('admin_logged_in'):
            flash('Please login to access the admin panel.', 'error')
            return redirect(url_for('admin_login', next=request.full_path))
        return f(*args, **kwargs)
    return decorated_function

@app.route('/admin')
@admin_required
def admin_dashboard():
    """Admin dashboard"""
    global CLUB_INFO, EVENTS, MEMBERS, GALLERY
    CLUB_INFO, EVENTS, MEMBERS, GALLERY = load_data()
    return render_template('admin/dashboard.html',
                         events_count=len(EVENTS),
                         members_count=len(MEMBERS),
                         gallery_count=len(GALLERY),
                         gallery=GALLERY)

@app.route('/admin/club-info', methods=['GET', 'POST'])
@admin_required
def admin_club_info():
    """Edit club information"""
    global CLUB_INFO, EVENTS, MEMBERS, GALLERY
    
    if request.method == 'POST':
        # Reload current club info
        CLUB_INFO, EVENTS, MEMBERS, GALLERY = load_data()
        
        # Handle logo upload
        logo_url = CLUB_INFO.get('logo', '/static/img/aicc-logo.webp')
        if 'logo_image' in request.files:
            file = request.files['logo_image']
            if file and file.filename and allowed_file(file.filename):
                # Delete old logo if it's in uploads folder
                delete_old_image(logo_url)
                
                filename = secure_filename(file.filename)
                timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
                filename = f"{timestamp}_{filename}"
                os.makedirs(app.config['UPLOAD_FOLDER'], exist_ok=True)
                filepath = os.path.join(app.config['UPLOAD_FOLDER'], filename)
                file.save(filepath)
                logo_url = f"/static/uploads/{filename}"
        
        # Process member_roles and member_years arrays from form
        member_roles = []
        member_years = []
        
        # Get roles from form (could be multiple inputs)
        role_data = request.form.get('member_roles_json')
        if role_data:
            try:
                member_roles = json.loads(role_data)
            except:
                member_roles = CLUB_INFO.get('member_roles', [])
        
        # Get years from form
        year_data = request.form.get('member_years_json')
        if year_data:
            try:
                member_years = json.loads(year_data)
            except:
                member_years = CLUB_INFO.get('member_years', [])
        
        data = {
            'name': request.form.get('name'),
            'short_name': request.form.get('short_name'),
            'tagline': request.form.get('tagline'),
            'description': request.form.get('description'),
            'college': request.form.get('college'),
            'department': request.form.get('department'),
            'address': request.form.get('address'),
            'logo': logo_url,
            'member_roles': member_roles,
            'member_years': member_years,
            'email': CLUB_INFO.get('email', ''),
            'linkedin': CLUB_INFO.get('linkedin', ''),
            'instagram': CLUB_INFO.get('instagram', ''),
            'email_config': {
                'MAIL_SERVER': request.form.get('mail_server', 'smtp.gmail.com'),
                'MAIL_PORT': int(request.form.get('mail_port', 587) or 587),
                'MAIL_USE_TLS': request.form.get('mail_use_tls') == 'true',
                'MAIL_USERNAME': request.form.get('mail_username', ''),
                'MAIL_PASSWORD': request.form.get('mail_password', ''),
                'MAIL_DEFAULT_SENDER': request.form.get('mail_default_sender', '')
            },
            'api_config': {
                'GROQ_API_KEY': request.form.get('groq_api_key', ''),
                'GROQ_MODEL': request.form.get('groq_model', 'llama-3.1-8b-instant'),
                'RAZORPAY_KEY_ID': request.form.get('razorpay_key_id', ''),
                'RAZORPAY_KEY_SECRET': request.form.get('razorpay_key_secret', '')
            },
            'faculty_coordinators': CLUB_INFO.get('faculty_coordinators', []),
            'secretaries': CLUB_INFO.get('secretaries', [])
        }
        
        with open(os.path.join(PROJECT_ROOT, 'data/club_info.json'), 'w') as f:
            json.dump(data, f, indent=4)
        
        # Reload data
        CLUB_INFO, EVENTS, MEMBERS, GALLERY = load_data()
        
        flash('Club information updated successfully!', 'success')
        return redirect(url_for('admin_club_info'))
    
    CLUB_INFO, EVENTS, MEMBERS, GALLERY = load_data()
    return render_template('admin/club_info.html', club_info=CLUB_INFO)

@app.route('/admin/events', methods=['GET'])
@admin_required
def admin_events():
    """View all events"""
    global CLUB_INFO, EVENTS, MEMBERS, GALLERY
    CLUB_INFO, EVENTS, MEMBERS, GALLERY = load_data()
    return render_template('admin/events.html', events=EVENTS)

@app.route('/admin/events/create', methods=['GET', 'POST'])
@admin_required
def admin_create_event():
    """Create a new event"""
    global CLUB_INFO, EVENTS, MEMBERS, GALLERY
    
    if request.method == 'POST':
        # Reload events from file
        with open(os.path.join(PROJECT_ROOT, 'data/events.json'), 'r') as f:
            events_data = json.load(f)
        
        # Handle both old array format and new object format
        if isinstance(events_data, list):
            events = events_data
            next_id = max([e.get('id', 0) for e in events], default=0) + 1
        else:
            events = events_data.get('events', [])
            next_id = events_data.get('next_id', 1)
        
        # Handle image upload
        image_url = ''
        if 'event_image' in request.files:
            file = request.files['event_image']
            if file and file.filename and allowed_file(file.filename):
                filename = secure_filename(file.filename)
                timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
                filename = f"{timestamp}_{filename}"
                os.makedirs(app.config['UPLOAD_FOLDER'], exist_ok=True)
                filepath = os.path.join(app.config['UPLOAD_FOLDER'], filename)
                file.save(filepath)
                image_url = f"/static/uploads/{filename}"
        
        # Add new event using next_id
        new_event = {
            'id': next_id,
            'name': request.form.get('name'),
            'date': request.form.get('date'),
            'time': request.form.get('time'),
            'location': request.form.get('location'),
            'description': request.form.get('description'),
            'how': request.form.get('how'),
            'status': request.form.get('status'),
            'image': image_url,
            'rules': request.form.get('rules', '').split('\n') if request.form.get('rules') else [],
            'coordinators': [],
            'show_in_events': request.form.get('show_in_events') == 'true'
        }
        
        # Handle registration settings
        registration_type = request.form.get('registration_type', 'none')
        new_event['registration_type'] = registration_type
        
        if registration_type == 'external':
            new_event['register_link'] = request.form.get('register_link', '#')
            new_event['template_id'] = None
        elif registration_type == 'internal':
            template_id = request.form.get('template_id')
            new_event['template_id'] = int(template_id) if template_id else None
            new_event['register_link'] = '#'
            
            # Create registration file for internal registration
            if new_event['template_id']:
                event_slug = re.sub(r'[^a-z0-9]+', '_', new_event['name'].lower()).strip('_')
                # Include event ID for uniqueness (same name events get different files)
                reg_filename = f"{event_slug}_{new_event['id']}_registrations.json"
                reg_file_path = os.path.join(PROJECT_ROOT, 'data', 'registrations', reg_filename)
                
                # Create registrations directory if it doesn't exist
                os.makedirs(os.path.dirname(reg_file_path), exist_ok=True)
                
                # Create empty registration file
                with open(reg_file_path, 'w') as f:
                    json.dump([], f)
                
                new_event['registration_file'] = f'data/registrations/{reg_filename}'
        else:
            new_event['register_link'] = '#'
            new_event['template_id'] = None
        
        # Add registration deadline if provided
        deadline_date = request.form.get('deadline_date')
        deadline_message = request.form.get('deadline_message')
        if deadline_date:
            new_event['registration_deadline'] = {
                'date': deadline_date,
                'message': deadline_message if deadline_message else 'Register now!'
            }
        
        events.append(new_event)
        
        # Save with incremented next_id
        with open(os.path.join(PROJECT_ROOT, 'data/events.json'), 'w') as f:
            json.dump({"next_id": next_id + 1, "events": events}, f, indent=4)
        
        # Reload data
        CLUB_INFO, EVENTS, MEMBERS, GALLERY = load_data()
        
        flash('Event created successfully!', 'success')
        return redirect(url_for('admin_events'))
    
    # Load form templates for the dropdown
    templates = []
    templates_file = os.path.join(PROJECT_ROOT, 'data', 'form_templates.json')
    try:
        if os.path.exists(templates_file):
            with open(templates_file, 'r') as f:
                templates = json.load(f)
    except Exception as e:
        logger.error(f"Error loading templates: {e}")
    
    return render_template('admin/create_event.html', forms=templates)

@app.route('/admin/events/<int:event_id>/delete', methods=['POST'])
@admin_required
def admin_delete_event(event_id):
    """Archive an event by marking it as completed (preserves registration data for attendance checks)"""
    global CLUB_INFO, EVENTS, MEMBERS, GALLERY
    
    events, next_id = load_events_file()
    
    # Find the event and mark as completed instead of deleting
    # This preserves registration data so students can still check their attendance
    event_to_archive = next((e for e in events if e.get('id') == event_id), None)
    if event_to_archive:
        event_to_archive['status'] = 'completed'
        event_to_archive['registration_type'] = 'none'  # Disable registration
        event_to_archive['allow_registration'] = False
    
    save_events_file(events, next_id)
    
    # Reload data
    CLUB_INFO, EVENTS, MEMBERS, GALLERY = load_data()
    
    flash('Event archived successfully! Registration data preserved for attendance checks.', 'success')
    return redirect(url_for('admin_events'))

@app.route('/admin/members', methods=['GET', 'POST'])
@admin_required
def admin_members():
    """Manage members"""
    global CLUB_INFO, EVENTS, MEMBERS, GALLERY
    
    if request.method == 'POST':
        with open(os.path.join(PROJECT_ROOT, 'data/members.json'), 'r') as f:
            members = json.load(f)
        
        # Handle image upload
        image_url = '/static/img/members/default.webp'
        if 'member_image' in request.files:
            file = request.files['member_image']
            if file and file.filename and allowed_file(file.filename):
                filename = secure_filename(file.filename)
                timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
                filename = f"{timestamp}_{filename}"
                os.makedirs(app.config['UPLOAD_FOLDER'], exist_ok=True)
                filepath = os.path.join(app.config['UPLOAD_FOLDER'], filename)
                file.save(filepath)
                image_url = f"/static/uploads/{filename}"
        
        new_member = {
            'name': request.form.get('name'),
            'role': request.form.get('role'),
            'year': request.form.get('year'),
            'domain': request.form.get('domain'),
            'image': image_url,
            'linkedin': request.form.get('linkedin'),
            'github': request.form.get('github')
        }
        
        members.append(new_member)
        
        # Sort members by role hierarchy and year before saving
        role_hierarchy = CLUB_INFO.get('member_roles', [])
        year_hierarchy = CLUB_INFO.get('member_years', [])
        if role_hierarchy:
            members = sort_members_by_role(members, role_hierarchy, year_hierarchy)
        
        with open(os.path.join(PROJECT_ROOT, 'data/members.json'), 'w') as f:
            json.dump(members, f, indent=4)
        
        # Reload data
        CLUB_INFO, EVENTS, MEMBERS, GALLERY = load_data()
        
        flash('Member added successfully!', 'success')
        return redirect(url_for('admin_members'))
    
    CLUB_INFO, EVENTS, MEMBERS, GALLERY = load_data()
    return render_template('admin/members.html', members=MEMBERS, club_info=CLUB_INFO)

@app.route('/admin/contact', methods=['GET', 'POST'])
@admin_required
def admin_contact():
    """Edit contact information"""
    global CLUB_INFO, EVENTS, MEMBERS, GALLERY
    
    if request.method == 'POST':
        # Load current club info and update contact fields
        CLUB_INFO, EVENTS, MEMBERS, GALLERY = load_data()
        
        CLUB_INFO['email'] = request.form.get('email')
        CLUB_INFO['instagram'] = request.form.get('instagram')
        CLUB_INFO['linkedin'] = request.form.get('linkedin')
        # Keep existing faculty_coordinators and secretaries
        
        with open(os.path.join(PROJECT_ROOT, 'data/club_info.json'), 'w') as f:
            json.dump(CLUB_INFO, f, indent=4)
        
        # Reload data
        CLUB_INFO, EVENTS, MEMBERS, GALLERY = load_data()
        
        flash('Contact information updated successfully!', 'success')
        return redirect(url_for('admin_contact'))
    
    CLUB_INFO, EVENTS, MEMBERS, GALLERY = load_data()
    return render_template('admin/contact.html', contact=CLUB_INFO)

# ========================================
# File Upload Routes
# ========================================

@app.route('/admin/upload', methods=['POST'])
@admin_required
def upload_file():
    """Handle file uploads"""
    if 'file' not in request.files:
        return jsonify({'error': 'No file provided'}), 400
    
    file = request.files['file']
    if file.filename == '':
        return jsonify({'error': 'No file selected'}), 400
    
    if file and allowed_file(file.filename):
        filename = secure_filename(file.filename)
        # Add timestamp to avoid conflicts
        timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
        filename = f"{timestamp}_{filename}"
        
        # Ensure upload folder exists
        os.makedirs(app.config['UPLOAD_FOLDER'], exist_ok=True)
        
        filepath = os.path.join(app.config['UPLOAD_FOLDER'], filename)
        file.save(filepath)
        
        # Return URL path
        url = f"/static/uploads/{filename}"
        return jsonify({'url': url}), 200
    
    return jsonify({'error': 'Invalid file type'}), 400

# ========================================
# Edit Routes
# ========================================

@app.route('/admin/events/<int:event_id>/edit', methods=['GET', 'POST'])
@admin_required
def admin_edit_event(event_id):
    """Edit an existing event"""
    global CLUB_INFO, EVENTS, MEMBERS, GALLERY
    
    events, next_id = load_events_file()
    
    event = next((e for e in events if e.get('id') == event_id), None)
    if not event:
        flash('Event not found!', 'error')
        return redirect(url_for('admin_events'))
    
    if request.method == 'POST':
        # Handle image upload
        image_url = event.get('image', '')
        if 'event_image' in request.files:
            file = request.files['event_image']
            if file and file.filename and allowed_file(file.filename):
                # Delete old image before uploading new one
                delete_old_image(image_url)
                
                filename = secure_filename(file.filename)
                timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
                filename = f"{timestamp}_{filename}"
                os.makedirs(app.config['UPLOAD_FOLDER'], exist_ok=True)
                filepath = os.path.join(app.config['UPLOAD_FOLDER'], filename)
                file.save(filepath)
                image_url = f"/static/uploads/{filename}"
        
        # Update event data
        event['name'] = request.form.get('name')
        event['date'] = request.form.get('date')
        event['time'] = request.form.get('time')
        event['location'] = request.form.get('location')
        event['description'] = request.form.get('description')
        event['how'] = request.form.get('how')
        event['status'] = request.form.get('status')
        event['image'] = image_url
        event['rules'] = request.form.get('rules', '').split('\n') if request.form.get('rules') else []
        event['show_in_events'] = request.form.get('show_in_events') == 'true'
        
        # Handle registration settings
        registration_type = request.form.get('registration_type', 'none')
        event['registration_type'] = registration_type
        
        if registration_type == 'external':
            event['register_link'] = request.form.get('register_link', '#')
            event['template_id'] = None
            # Clear registration_file for external registration
            if 'registration_file' in event:
                del event['registration_file']
        elif registration_type == 'internal':
            template_id = request.form.get('template_id')
            new_template_id = int(template_id) if template_id else None
            old_template_id = event.get('template_id')
            
            event['template_id'] = new_template_id
            event['register_link'] = '#'
            
            # Create/update registration file if template is set and no file exists
            if new_template_id and not event.get('registration_file'):
                # Generate registration filename based on event name and ID for uniqueness
                event_slug = re.sub(r'[^a-z0-9]+', '_', event['name'].lower()).strip('_')
                reg_filename = f"{event_slug}_{event['id']}_registrations.json"
                reg_file_path = os.path.join(PROJECT_ROOT, 'data', 'registrations', reg_filename)
                
                # Create registrations directory if it doesn't exist
                os.makedirs(os.path.dirname(reg_file_path), exist_ok=True)
                
                # Create empty registration file if it doesn't exist
                if not os.path.exists(reg_file_path):
                    with open(reg_file_path, 'w') as f:
                        json.dump([], f)
                
                # Update the registration_file path in event
                event['registration_file'] = f'data/registrations/{reg_filename}'
        else:
            event['register_link'] = '#'
            event['template_id'] = None
            # Clear registration_file if registration type is none
            if 'registration_file' in event:
                del event['registration_file']
        
        # Handle registration deadline
        deadline_date = request.form.get('deadline_date')
        deadline_message = request.form.get('deadline_message')
        if deadline_date:
            event['registration_deadline'] = {
                'date': deadline_date,
                'message': deadline_message if deadline_message else 'Register now!'
            }
        elif 'registration_deadline' in event:
            # Remove deadline if fields are empty
            del event['registration_deadline']
        
        save_events_file(events, next_id)
        
        # Reload data
        CLUB_INFO, EVENTS, MEMBERS, GALLERY = load_data()
        
        flash('Event updated successfully!', 'success')
        return redirect(url_for('admin_events'))
    
    # Load form templates for the dropdown
    templates = []
    templates_file = os.path.join(PROJECT_ROOT, 'data', 'form_templates.json')
    try:
        if os.path.exists(templates_file):
            with open(templates_file, 'r') as f:
                templates = json.load(f)
    except Exception as e:
        logger.error(f"Error loading templates: {e}")
    
    return render_template('admin/edit_event.html', event=event, forms=templates)

@app.route('/admin/events/<int:event_id>/delete-image', methods=['POST'])
@admin_required
def admin_delete_event_image(event_id):
    """Delete event image"""
    global CLUB_INFO, EVENTS, MEMBERS, GALLERY
    
    try:
        events, next_id = load_events_file()
        
        event = next((e for e in events if e.get('id') == event_id), None)
        if not event:
            return jsonify({'success': False, 'error': 'Event not found'}), 404
        
        # Delete the image file if it exists
        if event.get('image'):
            delete_old_image(event['image'])
            event['image'] = ''
            
            # Save updated events
            save_events_file(events, next_id)
            
            # Reload data
            CLUB_INFO, EVENTS, MEMBERS, GALLERY = load_data()
            
            return jsonify({'success': True})
        else:
            return jsonify({'success': False, 'error': 'No image to delete'}), 400
            
    except Exception as e:
        return jsonify({'success': False, 'error': str(e)}), 500

@app.route('/admin/members/<int:member_index>/edit', methods=['GET', 'POST'])
@admin_required
def admin_edit_member(member_index):
    """Edit an existing member"""
    global CLUB_INFO, EVENTS, MEMBERS, GALLERY
    
    with open(os.path.join(PROJECT_ROOT, 'data/members.json'), 'r') as f:
        members = json.load(f)
    
    if member_index >= len(members):
        flash('Member not found!', 'error')
        return redirect(url_for('admin_members'))
    
    member = members[member_index]
    
    if request.method == 'POST':
        # Handle image upload
        image_url = member.get('image', '')
        
        # Check if user wants to reset to default
        if request.form.get('reset_image') == 'true':
            # Delete old custom image if it exists
            if image_url and image_url != '/static/img/members/default.webp':
                delete_old_image(image_url)
            image_url = '/static/img/members/default.webp'
        elif 'member_image' in request.files:
            file = request.files['member_image']
            if file and file.filename and allowed_file(file.filename):
                # Delete old image before uploading new one
                delete_old_image(image_url)
                
                filename = secure_filename(file.filename)
                timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
                filename = f"{timestamp}_{filename}"
                os.makedirs(app.config['UPLOAD_FOLDER'], exist_ok=True)
                filepath = os.path.join(app.config['UPLOAD_FOLDER'], filename)
                file.save(filepath)
                image_url = f"/static/uploads/{filename}"
        
        # Update member data
        members[member_index] = {
            'name': request.form.get('name'),
            'role': request.form.get('role'),
            'year': request.form.get('year'),
            'domain': request.form.get('domain'),
            'image': image_url or '/static/img/members/default.webp',
            'linkedin': request.form.get('linkedin'),
            'github': request.form.get('github')
        }
        
        # Sort members by role hierarchy and year before saving
        role_hierarchy = CLUB_INFO.get('member_roles', [])
        year_hierarchy = CLUB_INFO.get('member_years', [])
        if role_hierarchy:
            members = sort_members_by_role(members, role_hierarchy, year_hierarchy)
        
        with open(os.path.join(PROJECT_ROOT, 'data/members.json'), 'w') as f:
            json.dump(members, f, indent=4)
        
        # Reload data
        CLUB_INFO, EVENTS, MEMBERS, GALLERY = load_data()
        
        flash('Member updated successfully!', 'success')
        return redirect(url_for('admin_members'))
    
    # Load club_info for role and year dropdowns
    CLUB_INFO, EVENTS, MEMBERS, GALLERY = load_data()
    return render_template('admin/edit_member.html', member=member, member_index=member_index, club_info=CLUB_INFO)

@app.route('/admin/members/<int:member_index>/delete', methods=['POST'])
@admin_required
def admin_delete_member(member_index):
    """Delete a member"""
    global CLUB_INFO, EVENTS, MEMBERS, GALLERY
    
    with open(os.path.join(PROJECT_ROOT, 'data/members.json'), 'r') as f:
        members = json.load(f)
    
    if member_index < len(members):
        # Delete member's image before removing from list
        member = members[member_index]
        delete_old_image(member.get('image', ''))
        
        members.pop(member_index)
        
        with open(os.path.join(PROJECT_ROOT, 'data/members.json'), 'w') as f:
            json.dump(members, f, indent=4)
        
        # Reload data
        CLUB_INFO, EVENTS, MEMBERS, GALLERY = load_data()
        
        flash('Member deleted successfully!', 'success')
    
    return redirect(url_for('admin_members'))

@app.route('/admin/gallery', methods=['GET', 'POST'])
@admin_required
def admin_gallery():
    """Manage gallery images"""
    global CLUB_INFO, EVENTS, MEMBERS, GALLERY
    
    if request.method == 'POST':
        if 'gallery_image' in request.files:
            file = request.files['gallery_image']
            if file and file.filename and allowed_file(file.filename):
                filename = secure_filename(file.filename)
                timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
                filename = f"{timestamp}_{filename}"
                os.makedirs(app.config['UPLOAD_FOLDER'], exist_ok=True)
                filepath = os.path.join(app.config['UPLOAD_FOLDER'], filename)
                file.save(filepath)
                
                # Add to gallery
                with open(os.path.join(PROJECT_ROOT, 'data/gallery.json'), 'r') as f:
                    gallery = json.load(f)
                
                new_image = {
                    'url': f"/static/uploads/{filename}",
                    'title': request.form.get('title', 'Gallery Image'),
                    'category': request.form.get('category', 'events')
                }
                
                gallery.append(new_image)
                
                with open(os.path.join(PROJECT_ROOT, 'data/gallery.json'), 'w') as f:
                    json.dump(gallery, f, indent=4)
                
                # Reload data
                CLUB_INFO, EVENTS, MEMBERS, GALLERY = load_data()
                
                flash('Image uploaded successfully!', 'success')
                return redirect(url_for('admin_gallery'))
    
    CLUB_INFO, EVENTS, MEMBERS, GALLERY = load_data()
    return render_template('admin/gallery.html', gallery=GALLERY)

@app.route('/admin/gallery/<int:image_index>/edit', methods=['GET', 'POST'])
@admin_required
def admin_edit_gallery_image(image_index):
    """Edit a gallery image"""
    global CLUB_INFO, EVENTS, MEMBERS, GALLERY
    
    with open(os.path.join(PROJECT_ROOT, 'data/gallery.json'), 'r') as f:
        gallery = json.load(f)
    
    if image_index >= len(gallery):
        flash('Image not found!', 'error')
        return redirect(url_for('admin_gallery'))
    
    image = gallery[image_index]
    
    if request.method == 'POST':
        # Update image details
        image['title'] = request.form.get('title')
        image['category'] = request.form.get('category', 'events')
        image['description'] = request.form.get('description', '')
        
        with open(os.path.join(PROJECT_ROOT, 'data/gallery.json'), 'w') as f:
            json.dump(gallery, f, indent=4)
        
        # Reload data
        CLUB_INFO, EVENTS, MEMBERS, GALLERY = load_data()
        
        flash('Image updated successfully!', 'success')
        return redirect(url_for('admin_gallery'))
    
    return render_template('admin/edit_gallery.html', image=image, image_index=image_index)

@app.route('/admin/gallery/<int:image_index>/delete', methods=['POST'])
@admin_required
def admin_delete_gallery_image(image_index):
    """Delete a gallery image"""
    global CLUB_INFO, EVENTS, MEMBERS, GALLERY
    
    with open(os.path.join(PROJECT_ROOT, 'data/gallery.json'), 'r') as f:
        gallery = json.load(f)
    
    if image_index < len(gallery):
        # Delete the image file before removing from gallery
        image = gallery[image_index]
        delete_old_image(image.get('url') or image.get('image', ''))
        
        gallery.pop(image_index)
        
        with open(os.path.join(PROJECT_ROOT, 'data/gallery.json'), 'w') as f:
            json.dump(gallery, f, indent=4)
        
        # Reload data
        CLUB_INFO, EVENTS, MEMBERS, GALLERY = load_data()
        
        flash('Image deleted successfully!', 'success')
    
    return redirect(url_for('admin_gallery'))

# ========================================
# REGISTRATION FORMS MANAGEMENT ROUTES
# ========================================

@app.route('/admin/form-templates')
@admin_required
def admin_registration_forms():
    """Admin page to manage form templates"""
    templates_file = os.path.join(PROJECT_ROOT, 'data', 'form_templates.json')
    templates = []
    
    try:
        if os.path.exists(templates_file):
            with open(templates_file, 'r') as f:
                templates = json.load(f)
    except Exception as e:
        flash('Error loading form templates.', 'error')
    
    return render_template('admin/registration_forms.html',
                         forms=templates)

@app.route('/admin/form-templates/create', methods=['GET', 'POST'])
@admin_required
def admin_create_registration_form():
    """Create a new form template"""
    if request.method == 'POST':
        try:
            # Load existing templates
            templates_file = os.path.join(PROJECT_ROOT, 'data', 'form_templates.json')
            templates = []
            if os.path.exists(templates_file):
                with open(templates_file, 'r') as f:
                    templates = json.load(f)
            
            # Generate unique ID
            max_id = max([t.get('id', 0) for t in templates], default=0)
            
            # Get form data with new participant-based structure
            template_data = {
                'id': max_id + 1,
                'name': request.form.get('name'),
                'description': request.form.get('description', ''),
                'min_participants': int(request.form.get('min_participants', 1)),
                'max_participants': int(request.form.get('max_participants', 1)),
                'custom_fields': json.loads(request.form.get('custom_fields', '[]')),
                'active': request.form.get('active') == 'true',
                'payment_enabled': request.form.get('payment_enabled') == 'true',
                'payment_amount': float(request.form.get('payment_amount', 0)) if request.form.get('payment_enabled') == 'true' else 0,
                'payment_description': request.form.get('payment_description', '') if request.form.get('payment_enabled') == 'true' else ''
            }
            
            # Add to templates list
            templates.append(template_data)
            
            # Save to file
            with open(templates_file, 'w') as f:
                json.dump(templates, f, indent=4)
            
            flash('Form template created successfully!', 'success')
            return redirect(url_for('admin_registration_forms'))
            
        except Exception as e:
            flash(f'Error creating form template: {str(e)}', 'error')
    
    return render_template('admin/create_registration_form.html')

@app.route('/admin/form-templates/<int:form_id>/edit', methods=['GET', 'POST'])
@admin_required
def admin_edit_registration_form(form_id):
    """Edit an existing form template"""
    templates_file = os.path.join(PROJECT_ROOT, 'data', 'form_templates.json')
    
    try:
        with open(templates_file, 'r') as f:
            templates = json.load(f)
    except:
        flash('Error loading form templates.', 'error')
        return redirect(url_for('admin_registration_forms'))
    
    # Find the template
    template_index = next((i for i, t in enumerate(templates) if t.get('id') == form_id), None)
    if template_index is None:
        flash('Template not found.', 'error')
        return redirect(url_for('admin_registration_forms'))
    
    if request.method == 'POST':
        try:
            # Update template data with new participant-based structure
            templates[template_index]['name'] = request.form.get('name')
            templates[template_index]['description'] = request.form.get('description', '')
            templates[template_index]['min_participants'] = int(request.form.get('min_participants', 1))
            templates[template_index]['max_participants'] = int(request.form.get('max_participants', 1))
            templates[template_index]['custom_fields'] = json.loads(request.form.get('custom_fields', '[]'))
            templates[template_index]['active'] = request.form.get('active') == 'true'
            templates[template_index]['payment_enabled'] = request.form.get('payment_enabled') == 'true'
            templates[template_index]['payment_amount'] = float(request.form.get('payment_amount', 0)) if request.form.get('payment_enabled') == 'true' else 0
            templates[template_index]['payment_description'] = request.form.get('payment_description', '') if request.form.get('payment_enabled') == 'true' else ''
            
            # Save to file
            with open(templates_file, 'w') as f:
                json.dump(templates, f, indent=4)
            
            flash('Form template updated successfully!', 'success')
            return redirect(url_for('admin_registration_forms'))
            
        except Exception as e:
            pass
            flash(f'Error updating form template: {str(e)}', 'error')
    
    return render_template('admin/edit_registration_form.html',
                         form=templates[template_index])

@app.route('/admin/form-templates/<int:form_id>/toggle', methods=['POST'])
@admin_required
def admin_toggle_form_status(form_id):
    """Toggle the active status of a form template"""
    templates_file = os.path.join(PROJECT_ROOT, 'data', 'form_templates.json')
    
    try:
        with open(templates_file, 'r') as f:
            templates = json.load(f)
        
        # Find the template and toggle its active status
        template = next((t for t in templates if t.get('id') == form_id), None)
        if template:
            template['active'] = not template.get('active', True)
            
            with open(templates_file, 'w') as f:
                json.dump(templates, f, indent=4)
            
            status = 'activated' if template['active'] else 'deactivated'
            flash(f'Form template {status} successfully!', 'success')
        else:
            flash('Template not found.', 'error')
            
    except Exception as e:
        flash('Error updating form template status.', 'error')
    
    return redirect(url_for('admin_registration_forms'))

@app.route('/admin/form-templates/<int:form_id>/delete', methods=['POST'])
@admin_required
def admin_delete_registration_form(form_id):
    """Delete a form template"""
    templates_file = os.path.join(PROJECT_ROOT, 'data', 'form_templates.json')
    
    try:
        with open(templates_file, 'r') as f:
            templates = json.load(f)
        
        # Find and remove the template
        template_index = next((i for i, t in enumerate(templates) if t.get('id') == form_id), None)
        if template_index is not None:
            templates.pop(template_index)
            
            with open(templates_file, 'w') as f:
                json.dump(templates, f, indent=4)
            
            flash('Form template deleted successfully!', 'success')
        else:
            flash('Template not found.', 'error')
            
    except Exception as e:
        flash('Error deleting form template.', 'error')
    
    return redirect(url_for('admin_registration_forms'))

@app.route('/admin/events/<int:event_id>/send-attendance-emails', methods=['POST'])
@admin_required
def admin_send_attendance_emails(event_id):
    """Send attendance verification emails to registrants"""
    global CLUB_INFO, EVENTS, MEMBERS, GALLERY
    CLUB_INFO, EVENTS, MEMBERS, GALLERY = load_data()
    
    try:
        data = request.get_json()
        filter_type = data.get('filter', 'marked')
        
        event = next((e for e in EVENTS if e.get('id') == event_id), None)
        if not event:
            return jsonify({'success': False, 'message': 'Event not found.'})
        
        # Load registrations
        registrations = []
        if event.get('registration_file'):
            reg_file = os.path.join(PROJECT_ROOT, event['registration_file'])
            if os.path.exists(reg_file):
                with open(reg_file, 'r') as f:
                    registrations = json.load(f)
        else:
            event_slug = slugify(event.get('name', ''))
            reg_file = os.path.join(PROJECT_ROOT, 'data', 'registrations', f'{event_slug}_registrations.json')
            if os.path.exists(reg_file):
                with open(reg_file, 'r') as f:
                    registrations = json.load(f)
        
        if not registrations:
            return jsonify({'success': False, 'message': 'No registrations found for this event.'})
        
        # Filter registrations based on filter type
        filtered_registrations = []
        for reg in registrations:
            status = reg.get('attendance_status', 'not_entered')
            if filter_type == 'all':
                filtered_registrations.append(reg)
            elif filter_type == 'marked' and status in ['entered', 'partially_present']:
                filtered_registrations.append(reg)
            elif filter_type == 'entered' and status == 'entered':
                filtered_registrations.append(reg)
            elif filter_type == 'partially_present' and status == 'partially_present':
                filtered_registrations.append(reg)
        
        if not filtered_registrations:
            return jsonify({'success': False, 'message': 'No registrations match the selected filter.'})
        
        # Send emails
        sent_count = 0
        failed_count = 0
        
        for reg in filtered_registrations:
            try:
                email = reg.get('submitter_email')
                if not email:
                    failed_count += 1
                    continue
                
                # Generate shareable link
                shareable_link = url_for('attendance_check',
                                        event_id=event_id,
                                        email=email,
                                        rid=reg.get('registration_id'),
                                        _external=True)
                
                # Generate QR code for the link
                qr = qrcode.QRCode(version=1, box_size=10, border=4)
                qr.add_data(shareable_link)
                qr.make(fit=True)
                qr_img = qr.make_image(fill_color="black", back_color="white")
                
                buffered = io.BytesIO()
                qr_img.save(buffered, format="PNG")
                qr_base64 = base64.b64encode(buffered.getvalue()).decode('utf-8')
                
                # Determine status text and styling
                status = reg.get('attendance_status', 'not_entered')
                if status == 'entered':
                    status_text = 'Fully Present'
                    status_color = '#10b981'
                    status_icon = '✓'
                elif status == 'partially_present':
                    status_text = 'Partially Present'
                    status_color = '#f59e0b'
                    status_icon = '⚠'
                else:
                    status_text = 'Attendance Not Marked'
                    status_color = '#6b7280'
                    status_icon = '○'
                
                # Get participant info
                participant_name = reg.get('name', reg.get('submitter_email', 'Participant'))
                
                # Build email HTML
                email_html = f"""
                <html>
                <body style="font-family: 'Segoe UI', Tahoma, Geneva, Verdana, sans-serif; margin: 0; padding: 20px; background-color: #f8f9fa;">
                    <div style="max-width: 600px; margin: 0 auto; background: white; border-radius: 16px; overflow: hidden; box-shadow: 0 4px 6px rgba(0, 0, 0, 0.1);">
                        <!-- Header -->
                        <div style="background: linear-gradient(135deg, #667eea 0%, #764ba2 100%); padding: 30px; text-align: center;">
                            <h1 style="color: white; margin: 0; font-size: 24px;">📋 Attendance Certificate</h1>
                            <p style="color: rgba(255,255,255,0.9); margin: 10px 0 0; font-size: 16px;">{html_escape(event.get('name', 'Event'))}</p>
                        </div>
                        
                        <!-- Status Badge -->
                        <div style="text-align: center; padding: 30px;">
                            <div style="display: inline-block; padding: 15px 30px; background: {status_color}20; border-radius: 50px; border: 2px solid {status_color};">
                                <span style="font-size: 24px; color: {status_color}; font-weight: bold;">{status_icon} {status_text}</span>
                            </div>
                        </div>
                        
                        <!-- Details -->
                        <div style="padding: 0 30px 30px;">
                            <h3 style="color: #1f2937; margin: 0 0 15px;">Hello {html_escape(participant_name)},</h3>
                            <p style="color: #4b5563; line-height: 1.6;">
                                Your attendance for <strong>{html_escape(event.get('name', 'the event'))}</strong> has been recorded.
                                Below are your details for verification:
                            </p>
                            
                            <table style="width: 100%; border-collapse: collapse; margin: 20px 0;">
                                <tr style="background: #f3f4f6;">
                                    <td style="padding: 12px 15px; border-bottom: 1px solid #e5e7eb; font-weight: 600; color: #374151;">Event</td>
                                    <td style="padding: 12px 15px; border-bottom: 1px solid #e5e7eb; color: #1f2937;">{html_escape(event.get('name', '-'))}</td>
                                </tr>
                                <tr>
                                    <td style="padding: 12px 15px; border-bottom: 1px solid #e5e7eb; font-weight: 600; color: #374151;">Date</td>
                                    <td style="padding: 12px 15px; border-bottom: 1px solid #e5e7eb; color: #1f2937;">{html_escape(event.get('date', '-'))}</td>
                                </tr>
                                <tr style="background: #f3f4f6;">
                                    <td style="padding: 12px 15px; border-bottom: 1px solid #e5e7eb; font-weight: 600; color: #374151;">Email</td>
                                    <td style="padding: 12px 15px; border-bottom: 1px solid #e5e7eb; color: #1f2937;">{html_escape(email)}</td>
                                </tr>
                                <tr>
                                    <td style="padding: 12px 15px; border-bottom: 1px solid #e5e7eb; font-weight: 600; color: #374151;">Registration ID</td>
                                    <td style="padding: 12px 15px; border-bottom: 1px solid #e5e7eb; color: #1f2937; font-family: monospace; font-size: 12px;">{html_escape(reg.get('registration_id', '-'))}</td>
                                </tr>
                                <tr style="background: #f3f4f6;">
                                    <td style="padding: 12px 15px; border-bottom: 1px solid #e5e7eb; font-weight: 600; color: #374151;">Attendance Status</td>
                                    <td style="padding: 12px 15px; border-bottom: 1px solid #e5e7eb; color: {status_color}; font-weight: bold;">{status_text}</td>
                                </tr>
                                {f'<tr><td style="padding: 12px 15px; border-bottom: 1px solid #e5e7eb; font-weight: 600; color: #374151;">Marked At</td><td style="padding: 12px 15px; border-bottom: 1px solid #e5e7eb; color: #1f2937;">{html_escape(reg.get("entry_time", "-"))}</td></tr>' if reg.get('entry_time') else ''}
                                {f'<tr style="background: #f3f4f6;"><td style="padding: 12px 15px; border-bottom: 1px solid #e5e7eb; font-weight: 600; color: #374151;">Comment</td><td style="padding: 12px 15px; border-bottom: 1px solid #e5e7eb; color: #f59e0b;">{html_escape(reg.get("attendance_comment", ""))}</td></tr>' if reg.get('attendance_comment') else ''}
                            </table>
                            
                            <!-- Verification Link -->
                            <div style="background: linear-gradient(135deg, #667eea20, #764ba220); border-radius: 12px; padding: 20px; margin: 20px 0; text-align: center;">
                                <p style="margin: 0 0 15px; color: #374151; font-weight: 600;">🔗 Shareable Verification Link</p>
                                <a href="{shareable_link}" style="display: inline-block; background: linear-gradient(135deg, #667eea, #764ba2); color: white; padding: 12px 25px; border-radius: 8px; text-decoration: none; font-weight: 600;">
                                    View Attendance Proof
                                </a>
                                <p style="margin: 15px 0 0; color: #6b7280; font-size: 12px;">
                                    Share this link with your faculty/college for verification
                                </p>
                            </div>
                            
                            <!-- QR Code -->
                            <div style="text-align: center; margin: 30px 0;">
                                <p style="color: #374151; margin: 0 0 10px; font-weight: 600;">📱 Scan QR to Verify</p>
                                <img src="cid:qr_code" alt="QR Code" style="width: 150px; height: 150px; border: 2px solid #e5e7eb; border-radius: 8px;">
                            </div>
                        </div>
                        
                        <!-- Footer -->
                        <div style="background: #f3f4f6; padding: 20px; text-align: center; border-top: 1px solid #e5e7eb;">
                            <p style="margin: 0; color: #6b7280; font-size: 12px;">
                                This is a computer-generated email from {html_escape(CLUB_INFO.get('name', 'AICC'))}.
                            </p>
                            <p style="margin: 5px 0 0; color: #9ca3af; font-size: 11px;">
                                © {datetime.now().year} {html_escape(CLUB_INFO.get('short_name', 'AICC'))}. All rights reserved.
                            </p>
                        </div>
                    </div>
                </body>
                </html>
                """
                
                # Create email message
                msg = Message(
                    subject=f"🎓 Attendance Certificate - {event.get('name', 'Event')}",
                    recipients=[email],
                    html=email_html
                )
                
                # Attach QR code as inline image
                qr_image_data = base64.b64decode(qr_base64)
                msg.attach(
                    'qr_code.png',
                    'image/png',
                    qr_image_data,
                    'inline',
                    headers={'Content-ID': '<qr_code>'}
                )
                
                mail.send(msg)
                sent_count += 1
                logging.info(f"Sent attendance email to {email} for event {event_id}")
                
            except Exception as e:
                logging.error(f"Failed to send attendance email to {reg.get('submitter_email')}: {e}")
                failed_count += 1
        
        message = f"Sent {sent_count} email(s) successfully."
        if failed_count > 0:
            message += f" {failed_count} failed."
        
        return jsonify({
            'success': True,
            'sent_count': sent_count,
            'failed_count': failed_count,
            'message': message
        })
        
    except Exception as e:
        logging.error(f"Error in send_attendance_emails: {e}")
        return jsonify({'success': False, 'message': str(e)})

@app.route('/admin/events/<int:event_id>/registrations')
@admin_required
def admin_view_registrations(event_id):
    """View registrations for a specific event"""
    global CLUB_INFO, EVENTS, MEMBERS, GALLERY
    CLUB_INFO, EVENTS, MEMBERS, GALLERY = load_data()
    
    event = next((e for e in EVENTS if e.get('id') == event_id), None)
    if not event:
        flash('Event not found.', 'error')
        return redirect(url_for('admin_events'))
    
    # Load form template if assigned
    template = None
    if event.get('template_id'):
        templates_file = os.path.join(PROJECT_ROOT, 'data', 'form_templates.json')
        try:
            with open(templates_file, 'r') as f:
                templates = json.load(f)
            template = next((t for t in templates if t.get('id') == event.get('template_id')), None)
        except:
            pass
    
    # Load registrations for this event
    registrations = []
    if event.get('registration_file'):
        reg_file = os.path.join(PROJECT_ROOT, event['registration_file'])
        if os.path.exists(reg_file):
            with open(reg_file, 'r') as f:
                registrations = json.load(f)
    else:
        # Fallback to old naming convention for backwards compatibility
        event_slug = slugify(event.get('name', ''))
        reg_file = os.path.join(PROJECT_ROOT, 'data', 'registrations', f'{event_slug}_registrations.json')
        if os.path.exists(reg_file):
            with open(reg_file, 'r') as f:
                registrations = json.load(f)
    
    return render_template('admin/view_registrations.html',
                         form=template,
                         event=event,
                         registrations=registrations)

@app.route('/admin/events/<int:event_id>/toggle-registration', methods=['POST'])
@admin_required
def admin_toggle_registration(event_id):
    """Toggle registration open/closed for an event"""
    global CLUB_INFO, EVENTS, MEMBERS, GALLERY
    
    try:
        events, next_id = load_events_file()
        
        event = next((e for e in events if e.get('id') == event_id), None)
        if not event:
            return jsonify({'success': False, 'message': 'Event not found'}), 404
        
        # Toggle the allow_registration field
        current_status = event.get('allow_registration', True)  # Default to True if not set
        event['allow_registration'] = not current_status
        
        save_events_file(events, next_id)
        
        # Reload data
        CLUB_INFO, EVENTS, MEMBERS, GALLERY = load_data()
        
        new_status = event['allow_registration']
        return jsonify({
            'success': True, 
            'allow_registration': new_status,
            'message': f'Registration {"enabled" if new_status else "disabled"} for {event.get("name")}'
        })
        
    except Exception as e:
        logger.error(f"Error toggling registration: {e}")
        return jsonify({'success': False, 'message': str(e)}), 500

@app.route('/admin/events/<int:event_id>/toggle-visibility', methods=['POST'])
@admin_required
def admin_toggle_visibility(event_id):
    """Toggle show_in_events for an event (show/hide from public Events page)"""
    global CLUB_INFO, EVENTS, MEMBERS, GALLERY
    
    try:
        events, next_id = load_events_file()
        
        event = next((e for e in events if e.get('id') == event_id), None)
        if not event:
            return jsonify({'success': False, 'message': 'Event not found'}), 404
        
        current = event.get('show_in_events', True)
        event['show_in_events'] = not current
        
        save_events_file(events, next_id)
        CLUB_INFO, EVENTS, MEMBERS, GALLERY = load_data()
        
        new_status = event['show_in_events']
        return jsonify({
            'success': True,
            'show_in_events': new_status,
            'message': f'Event {"shown" if new_status else "hidden"} on public Events page'
        })
        
    except Exception as e:
        logger.error(f"Error toggling visibility: {e}")
        return jsonify({'success': False, 'message': str(e)}), 500

@app.route('/admin/verify-entry')
@admin_required
def admin_verify_entry():
    """QR code entry verification page"""
    regid = request.args.get('regid')
    email = request.args.get('email')
    event_id = request.args.get('event_id')
    
    if not regid or not email or not event_id:
        flash('Invalid QR code. Missing parameters.', 'error')
        return redirect(url_for('admin_dashboard'))
    
    try:
        event_id = int(event_id)
    except (ValueError, TypeError):
        flash('Invalid event ID.', 'error')
        return redirect(url_for('admin_dashboard'))
    
    global CLUB_INFO, EVENTS, MEMBERS, GALLERY
    CLUB_INFO, EVENTS, MEMBERS, GALLERY = load_data()
    
    event = next((e for e in EVENTS if e.get('id') == event_id), None)
    if not event:
        flash('Event not found.', 'error')
        return redirect(url_for('admin_dashboard'))
    
    # Load registrations
    registrations = []
    reg_file_path = None
    if event.get('registration_file'):
        reg_file_path = os.path.join(PROJECT_ROOT, event['registration_file'])
        if os.path.exists(reg_file_path):
            with open(reg_file_path, 'r') as f:
                registrations = json.load(f)
    else:
        event_slug = slugify(event.get('name', ''))
        reg_file_path = os.path.join(PROJECT_ROOT, 'data', 'registrations', f'{event_slug}_registrations.json')
        if os.path.exists(reg_file_path):
            with open(reg_file_path, 'r') as f:
                registrations = json.load(f)
    
    # Find the registration
    registration = None
    for reg in registrations:
        if reg.get('registration_id') == regid and reg.get('submitter_email', '').lower() == email.lower():
            registration = reg
            break
    
    if not registration:
        flash('Registration not found or email does not match.', 'error')
        return redirect(url_for('admin_view_registrations', event_id=event_id))
    
    # Check attendance status
    attendance_status = registration.get('attendance_status', 'not_entered')
    already_marked = attendance_status in ['entered', 'partially_present']
    
    return render_template('admin/verify_entry.html',
                         event=event,
                         registration=registration,
                         already_marked=already_marked,
                         attendance_status=attendance_status)

@app.route('/admin/mark-entry', methods=['POST'])
@admin_required
def admin_mark_entry():
    """Mark attendee as entered (full, partial, or per-participant)"""
    regid = request.form.get('regid')
    email = request.form.get('email')
    event_id = request.form.get('event_id')
    attendance_type = request.form.get('attendance_type', 'full')  # 'full', 'partial', or 'participants'
    attendance_comment = request.form.get('attendance_comment', '').strip()
    participant_attendance_json = request.form.get('participant_attendance', '')
    
    if not regid or not email or not event_id:
        return jsonify({'error': 'Missing parameters'}), 400
    
    try:
        event_id = int(event_id)
    except (ValueError, TypeError):
        return jsonify({'error': 'Invalid event ID'}), 400
    
    global CLUB_INFO, EVENTS, MEMBERS, GALLERY
    CLUB_INFO, EVENTS, MEMBERS, GALLERY = load_data()
    
    event = next((e for e in EVENTS if e.get('id') == event_id), None)
    if not event:
        return jsonify({'error': 'Event not found'}), 404
    
    # Determine registration file path
    reg_file_path = None
    if event.get('registration_file'):
        reg_file_path = os.path.join(PROJECT_ROOT, event['registration_file'])
    else:
        event_slug = slugify(event.get('name', ''))
        reg_file_path = os.path.join(PROJECT_ROOT, 'data', 'registrations', f'{event_slug}_registrations.json')
    
    # Load registrations using safe read
    registrations = safe_json_read(reg_file_path)
    
    # Find and update the registration
    updated = False
    for reg in registrations:
        if reg.get('registration_id') == regid and reg.get('submitter_email', '').lower() == email.lower():
            
            # Handle participant-based attendance (checkboxes)
            if attendance_type == 'participants' and participant_attendance_json:
                try:
                    participant_attendance = json.loads(participant_attendance_json)
                    reg['participant_attendance'] = participant_attendance
                    
                    # Calculate overall attendance status
                    total = len(participant_attendance)
                    present = sum(1 for p in participant_attendance if p)
                    
                    if present == total:
                        reg['attendance_status'] = 'entered'
                    elif present > 0:
                        reg['attendance_status'] = 'partially_present'
                    else:
                        reg['attendance_status'] = 'not_entered'
                    
                    reg['attendance_comment'] = f'{present}/{total} participants present'
                    
                except (json.JSONDecodeError, TypeError) as e:
                    logger.error(f"Failed to parse participant_attendance: {e}")
                    return jsonify({'error': 'Invalid participant attendance data'}), 400
            else:
                # Legacy mode: full or partial attendance
                if attendance_type == 'partial':
                    reg['attendance_status'] = 'partially_present'
                else:
                    reg['attendance_status'] = 'entered'
                reg['attendance_comment'] = attendance_comment
            
            reg['entry_time'] = datetime.now().isoformat()
            reg['marked_by'] = session.get('admin_username', ADMIN_USERNAME)
            updated = True
            break
    
    if not updated:
        return jsonify({'error': 'Registration not found'}), 404
    
    # Save updated registrations using safe write
    try:
        safe_json_write(reg_file_path, registrations)
        return jsonify({'success': True, 'message': 'Entry marked successfully'}), 200
    except Exception as e:
        return jsonify({'error': 'Failed to save entry'}), 500

@app.route('/admin/events/<int:event_id>/registrations/export')
@admin_required
def admin_export_registrations(event_id):
    """Export registrations to Excel"""
    try:
        from openpyxl import Workbook
        from openpyxl.styles import Font, PatternFill, Alignment
        from flask import send_file
        from io import BytesIO
    except ImportError:
        flash('Please install openpyxl: pip install openpyxl', 'error')
        return redirect(url_for('admin_view_registrations', event_id=event_id))
    
    global CLUB_INFO, EVENTS, MEMBERS, GALLERY
    CLUB_INFO, EVENTS, MEMBERS, GALLERY = load_data()
    
    event = next((e for e in EVENTS if e.get('id') == event_id), None)
    if not event:
        flash('Event not found.', 'error')
        return redirect(url_for('admin_events'))
    
    # Load form template if assigned
    template = None
    if event.get('template_id'):
        templates_file = os.path.join(PROJECT_ROOT, 'data', 'form_templates.json')
        try:
            with open(templates_file, 'r') as f:
                templates = json.load(f)
            template = next((t for t in templates if t.get('id') == event.get('template_id')), None)
        except:
            pass
    
    if not template:
        flash('No form template found for this event.', 'error')
        return redirect(url_for('admin_view_registrations', event_id=event_id))
    
    # Load registrations
    registrations = []
    if event.get('registration_file'):
        reg_file = os.path.join(PROJECT_ROOT, event['registration_file'])
        if os.path.exists(reg_file):
            with open(reg_file, 'r') as f:
                registrations = json.load(f)
    else:
        # Fallback to old naming convention for backwards compatibility
        event_slug = slugify(event.get('name', ''))
        reg_file = os.path.join(PROJECT_ROOT, 'data', 'registrations', f'{event_slug}_registrations.json')
        if os.path.exists(reg_file):
            with open(reg_file, 'r') as f:
                registrations = json.load(f)
    
    if not registrations:
        flash('No registrations to export.', 'error')
        return redirect(url_for('admin_view_registrations', event_id=event_id))
    
    # Create Excel workbook
    wb = Workbook()
    ws = wb.active
    ws.title = 'Registrations'
    
    # Header style
    header_fill = PatternFill(start_color='4472C4', end_color='4472C4', fill_type='solid')
    header_font = Font(bold=True, color='FFFFFF')
    header_alignment = Alignment(horizontal='center', vertical='center')
    
    # Create headers
    headers = ['#', 'Timestamp', 'Submitter Email'] + [field.get('label') for field in template.get('fields', [])] + ['Payment Status', 'Attendance Status', 'Entry Time']
    for col_num, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_num, value=header)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = header_alignment
    
    # Add data rows
    for row_num, reg in enumerate(registrations, 2):
        ws.cell(row=row_num, column=1, value=reg.get('id', row_num - 1))
        ws.cell(row=row_num, column=2, value=reg.get('timestamp', ''))
        ws.cell(row=row_num, column=3, value=reg.get('submitter_email', '-'))
        
        for col_num, field in enumerate(template.get('fields', []), 4):
            value = reg.get(field.get('name'), '')
            ws.cell(row=row_num, column=col_num, value=str(value) if value else '-')
        
        # Add payment status, attendance status, and entry time
        payment_col = len(template.get('fields', [])) + 4  # After submitter email and form fields
        payment_status = reg.get('payment_status', 'not_required')
        ws.cell(row=row_num, column=payment_col, value=payment_status)
        
        attendance_col = payment_col + 1
        attendance_status = reg.get('attendance_status', 'not_entered')
        ws.cell(row=row_num, column=attendance_col, value=attendance_status)
        
        entry_time_col = payment_col + 2
        entry_time = reg.get('entry_time', '-')
        ws.cell(row=row_num, column=entry_time_col, value=entry_time)
    
    # Auto-adjust column widths
    for column in ws.columns:
        max_length = 0
        column_letter = column[0].column_letter
        for cell in column:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except:
                pass
        adjusted_width = min(max_length + 2, 50)
        ws.column_dimensions[column_letter].width = adjusted_width
    
    # Save to BytesIO
    output = BytesIO()
    wb.save(output)
    output.seek(0)
    
    # Generate filename
    filename = f"{event.get('name', 'event').replace(' ', '_')}_registrations.xlsx"
    
    return send_file(
        output,
        mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
        as_attachment=True,
        download_name=filename
    )

# ========================================
# Error Handlers
# ========================================

@app.errorhandler(404)
def page_not_found(e):
    """Custom 404 error page"""
    global CLUB_INFO, EVENTS, MEMBERS, GALLERY
    CLUB_INFO, EVENTS, MEMBERS, GALLERY = load_data()
    return render_template('404.html', club_info=CLUB_INFO, contact=CLUB_INFO), 404

if __name__ == '__main__':
    app.run(debug=True, host='0.0.0.0', port=5000)

